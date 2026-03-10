"""
build_master_index_v2.py
========================
KS TalentBridge — Enhanced Master Index Builder (v2)
Extends v1 (148 cols) with 53 new fields across 8 new categories:
  - Education structured (degree, field, school, year)
  - Work history (KSE start year, previous employers)
  - Largest project value
  - Additional roles (EIC, Office Engineer)
  - Software tools (MicroStation, AutoCAD, Revit, Procore, etc.)
  - Additional certifications (HAZWOPER, Traffic Control, Asphalt, SSPC, etc.)
  - Project types (16 flags: Bridge New/Rehab, Roadway, Water Main, Marine, etc.)
  - Additional agencies (TBTA, LIRR, Metro-North, NYCT, Nassau, Westchester, etc.)

Output: KSE_Master_Index_v2.xlsx  (TalentBridge root)
        KSE_Master_Index_v2.csv   (same data, plain text for ChatGPT / search)

Total columns: ~201
"""

import os, re, zipfile, json, sys, csv
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

try:
    import fitz
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False
    print("[WARN] PyMuPDF not installed — PDFs will be skipped. Run: pip install pymupdf")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl not installed — Excel output skipped. Run: pip install openpyxl")

# ── Config ────────────────────────────────────────────────────────────────────

_ROOT         = Path(__file__).resolve().parent.parent
RESUME_FOLDER = str(_ROOT / "data" / "source" / "Resume 08152025")
OUT_DIR       = str(_ROOT / "data" / "index")
RESUME_EXTS   = {".docx", ".pdf", ".rtf"}
W_NS          = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
NOW           = datetime.now().strftime("%Y-%m-%d %H:%M")

# ── Text Extraction ───────────────────────────────────────────────────────────

def _paras_from_xml(xml_bytes):
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return []
    out = []
    for p in root.iter(f"{W_NS}p"):
        t = "".join(n.text for n in p.iter(f"{W_NS}t") if n.text).strip()
        if t:
            out.append(t)
    return out

def extract_docx(path):
    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            parts = ["word/document.xml"]
            parts += sorted(n for n in names if re.match(r"word/(header|footer)\d*\.xml", n))
            parts += sorted(n for n in names if n.startswith("word/") and n.endswith(".xml")
                            and n not in parts and "theme" not in n and "settings" not in n
                            and "webSettings" not in n and "fontTable" not in n)
            paras = []
            for part in parts:
                try:
                    with z.open(part) as f:
                        paras.extend(_paras_from_xml(f.read()))
                except Exception:
                    pass
        seen, d = set(), []
        for p in paras:
            if p not in seen:
                d.append(p); seen.add(p)
        return "\n".join(d)
    except Exception:
        return ""

def extract_pdf(path):
    if not HAS_FITZ:
        return ""
    try:
        doc = fitz.open(path)
        pages = [p.get_text("text") for p in doc]
        doc.close()
        return "\n".join(pages)
    except Exception:
        return ""

def extract_rtf(path):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            raw = f.read()
        text = re.sub(r"\\[a-z]+\d*\s?", " ", raw)
        text = re.sub(r"[{}\\]", " ", text)
        return re.sub(r"\s+", " ", text).strip()
    except Exception:
        return ""

def extract_text(path):
    ext = Path(path).suffix.lower()
    if ext == ".docx": return extract_docx(path)
    if ext == ".pdf":  return extract_pdf(path)
    if ext == ".rtf":  return extract_rtf(path)
    return ""

# ── Name Cleanup ──────────────────────────────────────────────────────────────

def get_display_name(fname):
    stem = re.sub(r"\.(docx?|pdf|rtf)$", "", fname, flags=re.I)
    for pat in [
        r"^\d{6,8}[-_\s]*", r"^\d+\s*[-–]?\s*",
        r"^KSE[-_\s]+",
        r"^(?:MASTER|Master)[_\s]+(?:Resume[_\s]+[-_\s]*)?",
        r"^[-\s]*(?:MASTER|Master)[_\s]+(?:Resume[_\s]+)?[-\s]*",
        r"[-_\s]+KSE\b.*$", r"[-_\s]+NJDOT\b.*$", r"[-_\s]+Master\b.*$",
        r"[-_\s]*Resume\b.*$", r"[-_\s]*Format\b.*$", r"[-_\s]*SF\s*330\b.*$",
        r"[-_\s]*Design\s*Build\b.*$", r"[-_\s]*CADD\b.*$", r"[-_\s]*CVS\b.*$",
        r"[-_\s]*TCM\b.*$", r"[-_\s]*REV\b.*$", r"[-_\s]*draft\b.*$",
        r"\s*\d{1,2}[-./]\d{1,2}[-./]\d{2,4}.*$",
        r"[-_\s]+(?:Sr|Jr|Senior|Junior|Chief|Lead|Principal)\.?\s+\w.*$",
        r"[-_\s]+(?:Inspector|Engineer|Manager|Coordinator|Analyst|Designer).*$",
        r"[-_\s]+(?:Bridge|Highway|Structural|Civil|Environmental|Quality).*$",
        r"[-_\s]+(?:Focus|Insp|OE|TL|PM|SI|CCL|NII|HM)\b.*$",
        r"[_\-\s]+$",
    ]:
        stem = re.sub(pat, "", stem, flags=re.I)
    stem = re.sub(r"[_\-]+", " ", stem).strip().strip("., ")
    return stem if len(stem) > 1 else fname

# ── Helpers ───────────────────────────────────────────────────────────────────

def find(pattern, text, flags=re.I):
    m = re.search(pattern, text, flags)
    return m.group(1).strip() if m and m.lastindex and m.lastindex >= 1 else (m.group(0).strip() if m else "")

def has(pattern, text, flags=re.I):
    return "Yes" if re.search(pattern, text, flags) else ""

# ── Years of Experience ───────────────────────────────────────────────────────

def get_years_exp(text):
    m = re.search(r"(\d{1,2})\+?\s*(?:years?|yrs?)\s+(?:of\s+)?experience", text, re.I)
    if m:
        return m.group(1)
    m2 = re.search(r"TOTAL EXPERIENCE:\s*(\d+)\s*Months", text, re.I)
    if m2:
        return str(round(int(m2.group(1)) / 12))
    return ""

# ── Title Extraction ──────────────────────────────────────────────────────────

TITLE_RE = re.compile(
    r"\b((?:senior|junior|sr\.?|jr\.?|chief|lead|principal|resident|assistant|"
    r"assoc(?:iate)?|staff|project|licensed|registered|professional|deputy)\s+)?"
    r"(engineer|inspector|manager|director|coordinator|specialist|analyst|"
    r"technician|planner|architect|surveyor|superintendent|consultant|"
    r"officer|supervisor|designer|drafter|estimator|administrator)\b", re.I
)

def get_title(text):
    for line in text.splitlines()[:25]:
        line = line.strip()
        if 3 < len(line) < 100 and TITLE_RE.search(line):
            return line[:100]
    return ""

# ── PE License Extraction ─────────────────────────────────────────────────────

PE_STATES = ["NJ", "NY", "PA", "CT", "DE", "MD", "VA", "MA", "FL", "NC", "RI",
             "OH", "IL", "TX", "CA", "WY", "DC", "PR", "WV", "GA", "SC", "TN"]

def extract_pe_licenses(text):
    licenses = {}
    lines = text.splitlines()
    for state in PE_STATES:
        state_pat = re.compile(
            rf"(?:Professional\s+Engineer|P\.?E\.?)\s*[/\-\u2013\u2014:,]?\s*{state}\b[^,\n]{{0,5}}"
            rf"(?:[#\s]*|No\.?\s*|License\s*(?:No\.?)?\s*)?"
            rf"(PE[\s\-]?\d{{4,6}}[A-Z]?|\d{{4,6}}[A-Z]?|[A-Z]{{1,3}}\d{{4,6}}[A-Z]?)?",
            re.I
        )
        state_pat2 = re.compile(
            rf"{state}\s*\((?:[#\s]*|No\.?\s*|License\s*No\.?\s*)?"
            rf"(PE[\s\-]?\d{{4,6}}[A-Z]?|\d{{4,6}}[A-Z]?|[A-Z]{{2,3}}\d{{4,6}}[A-Z]?)",
            re.I
        )
        num = ""
        exp = ""
        for line in lines:
            if len(line) > 400:
                continue
            for pat in [state_pat, state_pat2]:
                m = pat.search(line)
                if m:
                    if m.lastindex and m.group(1):
                        num = m.group(1).strip()
                    exp_m = re.search(
                        r"[Ee]xp(?:ires?|\.?)?\s*[:\.]?\s*(\d{1,2}[/.\-]\d{1,2}[/.\-]?\d{2,4}|\d{4})",
                        line
                    )
                    if exp_m:
                        exp = exp_m.group(1)
                    break
            if num:
                break
        if not num:
            prose = re.compile(
                rf"licensed\s+(?:Professional\s+Engineer|P\.?E\.?)[^.\n]{{0,100}}{state}\b", re.I
            )
            if prose.search(text):
                num = "Yes (see resume)"
        if num:
            licenses[state] = {"num": num, "exp": exp}
    return licenses

# ── Other Licenses ────────────────────────────────────────────────────────────

def extract_other_licenses(text):
    result = {}
    pp = re.search(r"(?:Professional\s+Planner|P\.?P\.?)[^.\n]{0,40}NJ|NJ[^.\n]{0,30}Professional\s+Planner", text, re.I)
    if pp:
        num_m = re.search(r"(?:PP|Planner)[^.\n]{0,20}#?\s*(\d{4,8})", text, re.I)
        result["NJ_PP"] = num_m.group(1) if num_m else "Yes"
    if re.search(r"(?:Professional\s+Land\s+Surveyor|P\.?L\.?S\.?|Licensed\s+Land\s+Surveyor)", text, re.I):
        result["PLS"] = "Yes"
    if re.search(r"\bCFM\b|Certified\s+Floodplain\s+Manager", text, re.I):
        result["CFM"] = "Yes"
    if re.search(r"\bS\.?E\.?\b[^a-z]|Structural\s+Engineer\s+license", text, re.I):
        result["SE"] = "Yes"
    if re.search(r"\bR\.?A\.?\b[^a-z]|Registered\s+Architect", text, re.I):
        result["RA"] = "Yes"
    if re.search(r"\bCME\b|Certified\s+Municipal\s+Engineer", text, re.I):
        result["CME"] = "Yes"
    return result

# ── Certifications ────────────────────────────────────────────────────────────

def extract_certifications(text):
    certs = {}

    # NICET
    nicet = re.search(r"NICET[^.\n]{0,80}", text, re.I)
    if nicet:
        snippet = nicet.group(0)
        level_m = re.search(r"Level\s*(I{1,3}V?|\d)", snippet, re.I)
        certs["NICET"] = f"Level {level_m.group(1).upper()}" if level_m else "Yes"

    # ACI
    aci = re.search(r"\bACI\b[^.\n]{0,60}", text, re.I)
    if aci:
        snippet = aci.group(0)
        level_m = re.search(r"(?:Grade|Level)\s*(I{1,2}|\d)", snippet, re.I)
        certs["ACI"] = f"Grade {level_m.group(1).upper()}" if level_m else "Yes"

    certs["OSHA_30"] = "Yes" if re.search(r"\bOSHA[\s\-]?30\b", text, re.I) else ""
    certs["OSHA_10"] = "Yes" if re.search(r"\bOSHA[\s\-]?10\b", text, re.I) else ""

    # NACE
    nace = re.search(r"\bNACE\b[^.\n]{0,60}", text, re.I)
    if nace:
        snippet = nace.group(0)
        for lvl in ["3", "2", "1", "III", "II", "I"]:
            if re.search(rf"Level\s*{lvl}\b|Coating\s+Inspector\s*{lvl}", snippet, re.I):
                certs["NACE"] = f"Level {lvl}"
                break
        if "NACE" not in certs:
            certs["NACE"] = "Yes"

    certs["CWI"]  = "Yes" if re.search(r"\bCWI\b|Certified\s+Welding\s+Inspector", text, re.I) else ""
    certs["AWS"]  = "Yes" if re.search(r"\bAWS\b", text) else ""
    certs["PMP"]  = "Yes" if re.search(r"\bPMP\b|Project\s+Management\s+Professional", text, re.I) else ""
    certs["CCM"]  = "Yes" if re.search(r"\bCCM\b|Certified\s+Construction\s+Manager", text, re.I) else ""

    leed = re.search(r"\bLEED[^.\n]{0,30}", text, re.I)
    if leed:
        certs["LEED"] = leed.group(0)[:30].strip()

    certs["CHST"] = "Yes" if re.search(r"\bCHST\b", text, re.I) else ""
    certs["PTOE"] = "Yes" if re.search(r"\bPTOE\b", text, re.I) else ""
    certs["EIT"]  = "Yes" if re.search(r"\bE\.?I\.?T\.?\b", text, re.I) else ""
    certs["TWIC"] = "Yes" if re.search(r"\bTWIC\b", text, re.I) else ""
    certs["ICC"]  = "Yes" if re.search(r"\bICC\b|International\s+Code\s+Council", text, re.I) else ""
    certs["DBIA"] = "Yes" if re.search(r"\bDBIA\b|Design.Build\s+Institute", text, re.I) else ""

    # ── NEW v2 certifications ──────────────────────────────────────────────────
    certs["HAZWOPER"]            = "Yes" if re.search(r"\bhaz.?wo.?per\b", text, re.I) else ""
    certs["Traffic_Control_Coord"] = "Yes" if re.search(r"traffic\s+control\s+coord|(?<!\w)TCC(?!\w)|ATSSA", text, re.I) else ""
    certs["Asphalt_Tech"]        = "Yes" if re.search(r"asphalt.{0,20}technolog|NJSAT\b|SAT\s+cert|Asphalt\s+Paving\s+Technologist", text, re.I) else ""
    certs["SSPC"]                = "Yes" if re.search(r"\bSSPC\b", text) else ""
    certs["Confined_Space"]      = "Yes" if re.search(r"confined\s+space", text, re.I) else ""
    certs["Water_Main_Insp"]     = "Yes" if re.search(r"water\s+main.{0,30}certif|DDC.{0,20}water|NYCDDC.{0,20}water", text, re.I) else ""
    certs["NYCDOB"]              = "Yes" if re.search(r"\bNYCDOB\b|NYC\s+Dept\.?\s+of\s+Buildings|NYC\s+Department\s+of\s+Buildings", text, re.I) else ""

    return certs

# ── DOT / FHWA Training ───────────────────────────────────────────────────────

def extract_training(text):
    t = {}
    for lvl in ["1", "2", "3"]:
        key = f"PennDOT_RCI_L{lvl}"
        patterns = [
            rf"RCI\s+Level\s+{lvl}\b",
            rf"Roadway\s+Construction\s+Inspection[^.\n]{{0,40}}Level\s+{lvl}\b",
            rf"PennDOT[^.\n]{{0,40}}RCI[^.\n]{{0,20}}Level\s+{lvl}\b",
            rf"TA-TCI[^.\n]{{0,20}}{lvl}\b",
        ]
        t[key] = "Yes" if any(re.search(p, text, re.I) for p in patterns) else ""
    t["PennDOT_TCI"]      = "Yes" if re.search(r"\bTA-?TCI\b|\bTCIS\b|PennDOT.*TCI\b", text, re.I) else ""
    t["NBIS"]             = "Yes" if re.search(r"\bNBIS\b|National\s+Bridge\s+Inspection\s+Standard", text, re.I) else ""
    t["NHI_130055"]       = "Yes" if re.search(r"130055|Safety\s+Inspection\s+of\s+In.Service\s+Bridges", text, re.I) else ""
    t["NHI_130078_FC"]    = "Yes" if re.search(r"130078|Fracture\s+Critical\s+Inspection", text, re.I) else ""
    t["NHI_130053"]       = "Yes" if re.search(r"130053|Bridge\s+Inspection\s+Refresher", text, re.I) else ""
    t["NHI_135047_Scour"] = "Yes" if re.search(r"135047|Stream\s+Stability\s+and\s+Scour|Scour\s+at\s+Highway\s+Bridges", text, re.I) else ""
    t["NETTCP"]           = "Yes" if re.search(r"\bNETTCP\b", text, re.I) else ""
    t["Underwater_Insp"]  = "Yes" if re.search(r"underwater\s+(?:bridge\s+)?inspection|diving|diver", text, re.I) else ""
    return t

# ── Agency Experience (original 29 + 9 new) ───────────────────────────────────

AGENCIES = {
    "NJDOT":         [r"\bNJDOT\b", r"New\s+Jersey\s+Department\s+of\s+Transportation"],
    "PennDOT":       [r"\bPennDOT\b", r"\bPADOT\b", r"Pennsylvania\s+Department\s+of\s+Transportation"],
    "NYSDOT":        [r"\bNYSDOT\b", r"\bNYS\s+DOT\b", r"New\s+York\s+State\s+Dep(?:artment)?\s+of\s+Transportation"],
    "NYCDOT":        [r"\bNYCDOT\b", r"New\s+York\s+City\s+Dep(?:artment)?\s+of\s+Transportation"],
    "NJ_Transit":    [r"\bNJ\s+Transit\b", r"New\s+Jersey\s+Transit"],
    "PANYNJ":        [r"\bPANYNJ\b", r"Port\s+Authority\s+of\s+New\s+York", r"\bPA/NY\b"],
    "SEPTA":         [r"\bSEPTA\b", r"Southeastern\s+Pennsylvania\s+Transportation"],
    "Amtrak":        [r"\bAmtrak\b"],
    "Philadelphia_Streets": [r"(?:City\s+of\s+)?Philadelphia\s+(?:Department\s+of\s+)?Streets", r"Streets\s+Department"],
    "DelDOT":        [r"\bDelDOT\b", r"Delaware\s+Department\s+of\s+Transportation"],
    "FAA_Airport":   [r"\bFAA\b", r"Federal\s+Aviation", r"\bairport\b", r"\brunway\b", r"\btaxiway\b"],
    "NJTA_Turnpike": [r"\bNJTA\b", r"New\s+Jersey\s+Turnpike", r"Garden\s+State\s+Parkway", r"\bNJTP\b", r"\bGSP\b"],
    "PA_Turnpike":   [r"Pennsylvania\s+Turnpike", r"\bPA\s+Turnpike\b"],
    "DRPA":          [r"\bDRPA\b", r"Delaware\s+River\s+Port\s+Authority"],
    "DRJTBC":        [r"\bDRJTBC\b", r"Delaware\s+River\s+Joint\s+Toll\s+Bridge"],
    "MTA":           [r"\bMTA\b", r"Metropolitan\s+Transportation\s+Authority"],
    "NYC_Transit":   [r"\bNYCT\b", r"\bNYC\s+Transit\b", r"New\s+York\s+City\s+Transit"],
    "FHWA":          [r"\bFHWA\b", r"Federal\s+Highway\s+Administration"],
    "USACE":         [r"\bUSACE\b", r"Army\s+Corps\s+of\s+Engineers", r"Corps\s+of\s+Engineers"],
    "NJDEP":         [r"\bNJDEP\b", r"NJ\s+Dep(?:artment)?\s+of\s+Environmental"],
    "City_Newark":   [r"City\s+of\s+Newark\b", r"Newark[,\s]+NJ\b.*(?:project|contract|city)", r"Newark\s+(?:DPW|Dept|Department)"],
    "City_Philadelphia": [r"City\s+of\s+Philadelphia\b", r"Philadelphia\s+(?:Water\s+Dep|Streets|Parks|DDC|DPW)", r"\bPWD\b"],
    "FEMA":          [r"\bFEMA\b", r"Federal\s+Emergency\s+Management"],
    "County_Bridges_NJ": [r"(?:Essex|Hudson|Bergen|Passaic|Morris|Somerset|Union|Mercer|Monmouth|Ocean|Burlington|Camden|Gloucester|Atlantic|Cape\s+May|Cumberland|Salem|Hunterdon|Warren|Sussex)\s+Count(?:y|ies)?\s+(?:bridge|road|highway)", r"county\s+bridge.*NJ|NJ.*county\s+bridge"],
    "Water_Utilities_NJ": [r"NJ\s*(?:American\s+)?Water|water\s+main.*NJ|NJ.*water\s+main|New\s+Jersey\s+American\s+Water|Aqua\s+New\s+Jersey|NJAW"],
    "MassDOT":       [r"\bMassDOT\b", r"\bMBTA\b"],
    "ConnDOT":       [r"\bConnDOT\b", r"\bCTDOT\b", r"Connecticut\s+DOT"],
    "NYC_DDC":       [r"\bNYC\s+DDC\b", r"Department\s+of\s+Design\s+and\s+Construction"],
    "NYC_DEP":       [r"\bNYC\s+DEP\b", r"(?:NYC\s+)?Department\s+of\s+Environmental\s+Protection"],
    # ── NEW v2 agencies ──────────────────────────────────────────────────────
    "TBTA":           [r"\bTBTA\b", r"Triborough\s+Bridge(?:\s+and\s+Tunnel|\s+&\s+Tunnel)"],
    "MTA_LIRR":       [r"\bLIRR\b", r"Long\s+Island\s+Rail\s+Road"],
    "MTA_MetroNorth": [r"Metro.?North", r"MTA\s+MNR", r"\bMNR\b"],
    "NYCT_Subway":    [r"NYC\s*Transit|NYCT\b|MTA\s+NYCT|NYC\s+Subway"],
    "Nassau_County":  [r"Nassau\s+County(?:\s+DPW|\s+DOT|\s+bridge|\s+road)?"],
    "Suffolk_County": [r"Suffolk\s+County(?:\s+DPW|\s+DOT|\s+bridge|\s+road)?"],
    "Westchester_County": [r"Westchester\s+County(?:\s+DPW|\s+DOT|\s+bridge|\s+road)?"],
    "NJ_County_Govts":[r"(?:Bergen|Essex|Hudson|Middlesex|Morris|Union|Passaic|Somerset)\s+County\s+(?:Eng|DPW|Dept|Govern|road|bridge)"],
    "CTDOT":          [r"\bCTDOT\b", r"Connecticut\s+Department\s+of\s+Transportation"],
}

def extract_agencies(text):
    result = {}
    for agency, patterns in AGENCIES.items():
        result[agency] = "Yes" if any(re.search(p, text, re.I) for p in patterns) else ""
    return result

# ── Discipline Experience ─────────────────────────────────────────────────────

DISCIPLINES = {
    "Bridge_Inspection":    [r"bridge\s+inspection", r"\bNBIS\b", r"biennial\s+inspection", r"element.level\s+inspection", r"bridge\s+safety\s+inspection"],
    "Fracture_Critical":    [r"fracture\s+critical", r"\bFC\s+inspection\b", r"NHI\s*130078"],
    "Underwater_Inspection":[r"underwater\s+(?:bridge\s+)?inspection", r"(?:scuba\s+)?diver\b", r"diving\s+inspection"],
    "Bridge_Design":        [r"bridge\s+design", r"bridge\s+(?:rehabilitation|replacement|repair|widening)", r"structural\s+design.*bridge"],
    "Highway_Design":       [r"highway\s+design", r"roadway\s+design", r"geometric\s+design", r"alignment\s+design", r"plan.*profile.*design", r"road\s+design"],
    "Traffic_Signal_Design":[r"traffic\s+(?:engineering|signal|design)", r"signal\s+design", r"ITS\b", r"traffic\s+control", r"ATMS\b", r"signing\s+and\s+pavement"],
    "Hydrology_Hydraulics": [r"hydrology", r"hydraulics", r"HEC.RAS", r"HEC.HMS", r"stormwater\s+design", r"drainage\s+design", r"floodplain"],
    "Construction_Mgmt":    [r"construction\s+management", r"resident\s+engineer", r"construction\s+inspection", r"CM/CI\b", r"\bCM\b.*(?:services|contract)", r"construction\s+oversight"],
    "Special_Inspection":   [r"special\s+inspect(?:ion|or)", r"\bIBC\b", r"NYC\s+DOB", r"building\s+inspect(?:ion|or)", r"structural\s+observation"],
    "Geotechnical":         [r"geotechnical", r"soil\s+(?:boring|testing|investigation)", r"foundation\s+(?:design|analysis)", r"subsurface\s+investigation", r"bearing\s+capacity"],
    "Environmental":        [r"environmental\s+(?:assessment|impact|permit|review)", r"Phase\s+I|Phase\s+II", r"wetland\s+(?:permit|delineation)", r"NEPA\b", r"section\s+404"],
    "Steel_Repair_Rehab":   [r"structural\s+steel\s+(?:repair|rehab)", r"steel\s+(?:repair|rehabilitation|retrofit)", r"weld(?:ing)?\s+inspection", r"painting.*steel|steel.*painting"],
    "Coating_Painting_QA":  [r"(?:bridge\s+)?(?:coating|painting)\s+(?:inspection|QA|QC|quality)", r"NACE.*(?:inspect|QA|QC)", r"(?:paint|coating)\s+inspector", r"abrasive\s+blasting"],
    "Survey":               [r"land\s+survey", r"topographic\s+survey", r"route\s+survey", r"boundary\s+survey", r"\bGPS\s+survey\b", r"survey\s+manager"],
    "Geomatics":            [r"geomatics", r"LiDAR", r"bathymetric\s+survey"],
    "Water_Wastewater":     [r"water\s+main", r"sewer\s+(?:design|construction|inspection)", r"wastewater\s+treatment", r"pump\s+station", r"force\s+main"],
    "Utility_Coordination": [r"utility\s+(?:coordination|relocation|design)", r"subsurface\s+utility", r"\bSUE\b"],
    "Marine_Waterfront":    [r"marine\s+(?:structure|inspection)", r"\bpier\s+inspection\b", r"bulkhead\s+(?:inspection|design)", r"waterfront\s+(?:structure|inspection)", r"dock\s+inspection"],
    "Rail_Transit":         [r"railroad\s+(?:inspection|design|crossing)", r"commuter\s+rail", r"light\s+rail", r"transit\s+(?:station|facility)", r"grade\s+crossing", r"rail\s+corridor"],
    "Airport":              [r"airport\s+(?:pavement|runway|taxiway)", r"runway\s+(?:design|inspection)", r"\bFAA\s+project\b"],
    "Culvert_Inspection":   [r"culvert\s+inspection", r"large\s+culvert\s+inspection", r"culvert\s+inspector"],
    "Design_Build":         [r"design.build", r"\bDBB\b", r"DBIA\b"],
    "QA_QC_Program":        [r"QA/QC\s+(?:program|plan|manager|manager)", r"quality\s+assurance.*quality\s+control", r"quality\s+management\s+(?:plan|program)"],
}

def extract_disciplines(text):
    result = {}
    for disc, patterns in DISCIPLINES.items():
        result[disc] = "Yes" if any(re.search(p, text, re.I) for p in patterns) else ""
    return result

# ── Leadership Roles (original 10 + 2 new) ───────────────────────────────────

ROLES = {
    "Project_Manager":      [r"\bProject\s+Manager\b", r"\bPM\b[^A-Z]", r"managing\s+projects", r"project\s+management\s+(?:role|experience|responsibilities)"],
    "Resident_Engineer":    [r"Resident\s+Engineer\b", r"\bRE\b\s+(?:for|on|at)", r"construction\s+resident"],
    "Team_Leader":          [r"\bTeam\s+Leader\b", r"\bTL\b\s+(?:for|on|bridge)", r"lead(?:ing)?\s+(?:a\s+)?(?:inspection\s+)?team"],
    "Chief_Inspector":      [r"Chief\s+Inspector\b", r"Senior\s+Inspector.*(?:in\s+charge|supervising)", r"Inspector.*in\s+Charge\b"],
    "Inspector_in_Charge":  [r"Inspector\s+in\s+Charge\b", r"\bIIC\b"],
    "Lead_Designer":        [r"Lead\s+(?:Design(?:er)?|Engineer)\b", r"Design\s+(?:Team\s+)?Lead(?:er)?\b"],
    "QA_QC_Manager":        [r"QA/QC\s+Manager\b", r"Quality\s+(?:Assurance\s+)?Manager\b", r"Quality\s+Control\s+Manager\b"],
    "Survey_Manager":       [r"Survey\s+Manager\b", r"Survey\s+(?:Team\s+)?Lead(?:er)?\b"],
    "Department_Manager":   [r"Department\s+Manager\b", r"Office\s+Manager\b", r"Group\s+Manager\b"],
    "Principal_in_Charge":  [r"Principal.in.Charge\b", r"Principal\s+(?:Engineer|Manager)\b"],
    # ── NEW v2 roles ─────────────────────────────────────────────────────────
    "EIC_Engineer_in_Charge": [r"\bEngineer.in.Charge\b", r"\bEIC\b\s+(?:for|on|at|while)", r"served\s+as\s+(?:an?\s+)?EIC\b"],
    "Office_Engineer":        [r"\bOffice\s+Engineer\b", r"\bOE\b\s+(?:for|on|at|training)"],
}

def extract_roles(text):
    result = {}
    for role, patterns in ROLES.items():
        result[role] = "Yes" if any(re.search(p, text, re.I) for p in patterns) else ""
    return result

# ── Special Flags ─────────────────────────────────────────────────────────────

def extract_flags(text):
    flags = {}
    flags["DBE_MBE_WBE"]     = "Yes" if re.search(r"\b(?:DBE|MBE|WBE|SBE|LBE|SDVOB)\b", text, re.I) else ""
    agency_count = sum(1 for pats in AGENCIES.values() if any(re.search(p, text, re.I) for p in pats))
    flags["Multi_Agency"]     = "Yes" if agency_count >= 3 else ""
    flags["Multi_Million_PM"] = "Yes" if re.search(r"\$\s*\d+[\.,]\d*\s*[Mm]illion|\$\s*\d{2,}[Mm]\b", text, re.I) else ""
    flags["Design_Build"]     = "Yes" if re.search(r"design.build", text, re.I) else ""
    return flags

# ── Education (original summary + NEW structured fields) ─────────────────────

def extract_education(text):
    m = re.search(
        r"\b(B\.?S\.?|M\.?S\.?|Ph\.?D\.?|B\.?E\.?|M\.?E\.?|B\.?Eng\.?|M\.?Eng\.?|MBA)\b[^.\n]{0,80}"
        r"(?:Civil|Structural|Environmental|Transportation|Construction|Engineering|Architecture|Science|Planning)",
        text, re.I
    )
    return m.group(0)[:100].strip() if m else ""

def extract_education_structured(text):
    """Returns dict with degree_type, edu_field, edu_school, edu_year."""
    result = {"edu_degree": "", "edu_field": "", "edu_school": "", "edu_year": ""}

    # Look for patterns like: "BS/Civil Engineering, University Name, 1997"
    # or "Bachelor of Science, Civil Engineering - Some University - 2005"
    degree_map = {
        r"Ph\.?D\.?|Doctor\s+of\s+Philosophy": "PhD",
        r"M\.?S\.?|Master\s+of\s+Science|Masters?\s+Degree": "MS",
        r"M\.?Eng\.?|Master\s+of\s+Engineering": "MEng",
        r"MBA|Master\s+of\s+Business": "MBA",
        r"M\.?Arch\.?|Master\s+of\s+Architecture": "MArch",
        r"B\.?S\.?|Bachelor\s+of\s+Science|B\.?Sc\.?": "BS",
        r"B\.?E\.?|Bachelor\s+of\s+Engineering": "BE",
        r"B\.?Arch\.?|Bachelor\s+of\s+Architecture": "BArch",
        r"AAS|Associate": "AAS",
    }

    field_patterns = [
        r"Civil\s+Engineering", r"Structural\s+Engineering", r"Environmental\s+Engineering",
        r"Transportation\s+Engineering", r"Construction\s+(?:Engineering|Management|Technology)",
        r"Mechanical\s+Engineering", r"Electrical\s+Engineering", r"Computer\s+(?:Science|Engineering)",
        r"Architecture", r"Urban\s+Planning", r"Geology", r"Mathematics", r"Physics",
        r"Business\s+Administration", r"Marketing", r"Engineering\s+Technology",
    ]

    # Find education section
    edu_section = ""
    m_sec = re.search(r"EDUCATION[^\n]*\n((?:[^\n]{0,200}\n){1,8})", text, re.I)
    if m_sec:
        edu_section = m_sec.group(1)
    else:
        edu_section = text[:2000]

    # Extract degree type
    for pat, label in degree_map.items():
        if re.search(pat, edu_section, re.I):
            result["edu_degree"] = label
            break

    # Extract field
    for fpat in field_patterns:
        if re.search(fpat, edu_section, re.I):
            m_field = re.search(fpat, edu_section, re.I)
            result["edu_field"] = m_field.group(0).strip()
            break

    # Extract school name
    school_pat = re.compile(
        r"(?:University|College|Institute|School|Academy)\s+of\s+[\w\s,]+|"
        r"[\w\s]+(?:University|College|Institute|Polytechnic|School)\b",
        re.I
    )
    m_school = school_pat.search(edu_section)
    if m_school:
        result["edu_school"] = m_school.group(0).strip()[:80]

    # Extract year
    m_year = re.search(r"(?:19|20)\d{2}", edu_section)
    if m_year:
        result["edu_year"] = m_year.group(0)

    return result

# ── Work History ──────────────────────────────────────────────────────────────

def extract_work_history(text):
    """Extract KSE start year, previous employers list, and employer count."""
    result = {"kse_start_year": "", "prev_employers": "", "num_prev_employers": ""}

    # Look for WORK HISTORY section
    wh_match = re.search(r"WORK\s+HISTORY[^\n]*\n((?:[^\n]{0,200}\n){1,15})", text, re.I)
    if not wh_match:
        # Try "EMPLOYMENT HISTORY" or similar
        wh_match = re.search(r"EMPLOYMENT\s+(?:HISTORY|EXPERIENCE)[^\n]*\n((?:[^\n]{0,200}\n){1,15})", text, re.I)

    if wh_match:
        wh_text = wh_match.group(1)

        # KSE start year — look for "KS Engineers" with a year
        kse_m = re.search(r"KS\s+Engineers?[^(\n]{0,30}\(?(\d{4})\s*[-–]", wh_text, re.I)
        if kse_m:
            result["kse_start_year"] = kse_m.group(1)

        # Extract employer names — lines with a year pattern
        employers = []
        for line in wh_text.splitlines():
            line = line.strip()
            if len(line) < 5:
                continue
            # Skip KSE lines
            if re.search(r"KS\s+Engineers", line, re.I):
                continue
            # Lines with a year (employer entries typically have dates)
            if re.search(r"(?:19|20)\d{2}", line):
                # Clean up the employer name — remove date part
                emp = re.sub(r"\s*[\(\[]?\s*(?:19|20)\d{2}\s*[-–]\s*(?:(?:19|20)\d{2}|[Pp]resent)\s*[\)\]]?\s*$", "", line)
                emp = re.sub(r"\s*\(?\s*(?:19|20)\d{2}\s*[-–].*$", "", emp).strip()
                emp = re.sub(r"[,;]\s*$", "", emp).strip()
                if emp and len(emp) > 3 and not re.search(r"KS\s+Engineers", emp, re.I):
                    employers.append(emp)

        if employers:
            result["prev_employers"] = "; ".join(employers[:8])
            result["num_prev_employers"] = str(len(employers[:8]))

    return result

# ── Largest Project Value ─────────────────────────────────────────────────────

def extract_largest_project(text):
    """Find the largest dollar value project mentioned (in millions)."""
    values = []
    # Match $XXX million, $XXX.X million, $XXXm, $XXXm
    for m in re.finditer(
        r"\$\s*([\d,]+(?:\.\d+)?)\s*(?:million|M\b)",
        text, re.I
    ):
        try:
            val = float(m.group(1).replace(",", ""))
            values.append(val)
        except ValueError:
            pass
    # Also match $X,XXX,XXX style (>$1M)
    for m in re.finditer(r"\$([\d,]{7,})", text):
        try:
            val = float(m.group(1).replace(",", "")) / 1_000_000
            if val >= 1:
                values.append(val)
        except ValueError:
            pass
    if values:
        return f"${max(values):.1f}M"
    return ""

# ── Software Tools (NEW v2) ───────────────────────────────────────────────────

SOFTWARE_TOOLS = {
    "MicroStation":  [r"\bMicroStation\b"],
    "AutoCAD":       [r"\bAutoCAD\b"],
    "ProjectWise":   [r"\bProjectWise\b"],
    "SiteManager":   [r"\bSite\s*Manager\b"],
    "HEC_RAS":       [r"\bHEC.?RAS\b"],
    "GIS_ArcGIS":    [r"\bArcGIS\b", r"\bGIS\b"],
    "Procore":       [r"\bProcore\b"],
    "Primavera_P6":  [r"\bPrimavera\b", r"\bP6\b\s+(?:scheduling|software|project)"],
    "STAAD_SAP2000": [r"\bSTAAD\b", r"\bSAP2000\b"],
    "Revit":         [r"\bRevit\b"],
    "MATLAB":        [r"\bMATLAB\b"],
}

def extract_software(text):
    result = {}
    for tool, patterns in SOFTWARE_TOOLS.items():
        result[tool] = "Yes" if any(re.search(p, text, re.I) for p in patterns) else ""
    return result

# ── Project Types (NEW v2) ────────────────────────────────────────────────────

PROJECT_TYPES = {
    "Bridge_New_Construction":  [r"new\s+bridge\s+construction", r"bridge\s+replacement", r"replacement\s+of.{0,30}bridge", r"new\s+bridge\s+over"],
    "Bridge_Rehabilitation":    [r"bridge\s+(?:rehab|repair|restoration|widening)", r"rehabilitation\s+of.{0,30}bridge"],
    "Roadway_Construction":     [r"roadway\s+construction", r"highway\s+construction", r"road\s+construction"],
    "Roadway_Resurfacing":      [r"resurfac(?:ing)?", r"milling\s+and\s+inlay", r"mill\s+and\s+(?:inlay|pave|paving)", r"\bHMA\b", r"asphalt\s+(?:paving|overlay)"],
    "Drainage_Stormwater":      [r"drainage\s+(?:improvement|design|construction)", r"stormwater", r"culvert\s+(?:replacement|installation|construction)"],
    "Water_Main_Sewer":         [r"water\s+main", r"water\s+distribution", r"sewer\s+(?:line|construction|replacement)", r"force\s+main"],
    "Rail_Transit_Construction":[r"rail\s+(?:yard|line|track)\s+construction", r"track\s+construction", r"railroad\s+construction", r"transit\s+(?:facility|station)\s+construction"],
    "Traffic_Signal":           [r"traffic\s+signal\s+(?:construction|installation|design)", r"new\s+signal\s+system"],
    "ADA_Accessibility":        [r"ADA\s+(?:improvements?|compliance|ramps?|accessible)", r"accessibility\s+improvements?", r"handicap\s+ramp"],
    "Streetscape":              [r"streetscape", r"pedestrian\s+improvements?", r"pedestrian\s+plaza"],
    "Marine_Waterfront":        [r"marine\s+(?:structure|construction)", r"bulkhead\s+construction", r"pier\s+(?:construction|replacement)", r"fender\s+(?:system|replacement)", r"waterfront\s+construction"],
    "Movable_Bridge":           [r"bascule\s+bridge", r"movable\s+bridge", r"swing\s+bridge", r"draw\s*bridge", r"drawbridge"],
    "Suspension_CableStayed":   [r"suspension\s+bridge", r"cable.stayed\s+bridge"],
    "Historic_Structure":       [r"historic\s+bridge", r"historic\s+(?:structure|building|preserv)", r"preservation\s+of.{0,30}historic"],
    "Emergency_Repair":         [r"emergency\s+repair", r"emergency\s+response", r"emergency\s+(?:contract|project)", r"storm\s+damage\s+repair"],
    "Federal_Facility":         [r"federal\s+facilit", r"military\s+base", r"\bNASA\b", r"(?:USACE|Army\s+Corps).{0,30}construction"],
}

def extract_project_types(text):
    result = {}
    for ptype, patterns in PROJECT_TYPES.items():
        result[ptype] = "Yes" if any(re.search(p, text, re.I) for p in patterns) else ""
    return result

# ── Main Record Builder ───────────────────────────────────────────────────────

def build_record(fpath):
    fname = os.path.basename(fpath)
    text  = extract_text(fpath)

    if not text.strip():
        return None, "no_text"

    name      = get_display_name(fname)
    title     = get_title(text)
    years_exp = get_years_exp(text)
    education = extract_education(text)
    edu_str   = extract_education_structured(text)
    work_hist = extract_work_history(text)
    largest_proj = extract_largest_project(text)
    pe        = extract_pe_licenses(text)
    other_lic = extract_other_licenses(text)
    certs     = extract_certifications(text)
    training  = extract_training(text)
    agencies  = extract_agencies(text)
    discs     = extract_disciplines(text)
    roles     = extract_roles(text)
    flags     = extract_flags(text)
    software  = extract_software(text)
    proj_types= extract_project_types(text)

    return {
        # ── Identity ──────────────────────────────────────────────────────────
        "display_name":   name,
        "file_name":      fname,
        "title":          title,
        "years_exp":      years_exp,
        "education":      education,
        # ── Education Structured (NEW v2) ─────────────────────────────────────
        "edu_degree":     edu_str["edu_degree"],
        "edu_field":      edu_str["edu_field"],
        "edu_school":     edu_str["edu_school"],
        "edu_year":       edu_str["edu_year"],
        # ── Work History (NEW v2) ─────────────────────────────────────────────
        "kse_start_year":     work_hist["kse_start_year"],
        "prev_employers":     work_hist["prev_employers"],
        "num_prev_employers": work_hist["num_prev_employers"],
        # ── Largest Project (NEW v2) ──────────────────────────────────────────
        "largest_project_value": largest_proj,
        # ── PE Licenses ───────────────────────────────────────────────────────
        **{f"PE_{s}_num": pe.get(s, {}).get("num", "") for s in PE_STATES},
        **{f"PE_{s}_exp": pe.get(s, {}).get("exp", "") for s in PE_STATES},
        # ── Other Licenses ────────────────────────────────────────────────────
        "NJ_PP": other_lic.get("NJ_PP", ""),
        "PLS":   other_lic.get("PLS", ""),
        "CFM":   other_lic.get("CFM", ""),
        "SE":    other_lic.get("SE", ""),
        "RA":    other_lic.get("RA", ""),
        "CME":   other_lic.get("CME", ""),
        # ── Certifications ────────────────────────────────────────────────────
        **{f"CERT_{k}": v for k, v in certs.items()},
        # ── Training ──────────────────────────────────────────────────────────
        **training,
        # ── Agencies ──────────────────────────────────────────────────────────
        **{f"AGY_{k}": v for k, v in agencies.items()},
        # ── Disciplines ───────────────────────────────────────────────────────
        **{f"DISC_{k}": v for k, v in discs.items()},
        # ── Roles ─────────────────────────────────────────────────────────────
        **{f"ROLE_{k}": v for k, v in roles.items()},
        # ── Flags ─────────────────────────────────────────────────────────────
        **{f"FLAG_{k}": v for k, v in flags.items()},
        # ── Software Tools (NEW v2) ───────────────────────────────────────────
        **{f"SW_{k}": v for k, v in software.items()},
        # ── Project Types (NEW v2) ────────────────────────────────────────────
        **{f"PROJ_{k}": v for k, v in proj_types.items()},
    }, None

# ── Deduplicate ───────────────────────────────────────────────────────────────

def merge_records(records):
    merged = {}
    for rec in records:
        name_key = re.sub(r"\s+", "", rec["display_name"].lower())
        if name_key not in merged:
            merged[name_key] = rec.copy()
        else:
            existing = merged[name_key]
            for k, v in rec.items():
                if not existing.get(k) and v:
                    existing[k] = v
    return list(merged.values())

# ── Excel Output ──────────────────────────────────────────────────────────────

def section_color(key):
    if key in ("display_name","file_name","title","years_exp","education",
               "edu_degree","edu_field","edu_school","edu_year",
               "kse_start_year","prev_employers","num_prev_employers","largest_project_value"):
        return "1F4E79"
    if key.startswith("PE_") and "_num" in key: return "375623"
    if key.startswith("PE_") and "_exp" in key: return "4F8A10"
    if key in ("NJ_PP","PLS","CFM","SE","RA","CME"): return "375623"
    if key.startswith("CERT_"): return "7030A0"
    if key.startswith(("PennDOT","NBIS","NHI","NETTCP","Underwater")): return "833C00"
    if key.startswith("AGY_"): return "0070C0"
    if key.startswith("DISC_"): return "C55A11"
    if key.startswith("ROLE_"): return "843C0C"
    if key.startswith("FLAG_"): return "7F7F7F"
    if key.startswith("SW_"):   return "006064"
    if key.startswith("PROJ_"): return "BF360C"
    return "1F4E79"

def write_excel(records, out_path):
    if not HAS_OPENPYXL:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Index v2"

    if not records:
        wb.save(out_path)
        return

    all_keys = list(records[0].keys())

    for c, key in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=c, value=key.replace("_", " "))
        color = section_color(key)
        cell.fill      = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    fill_a = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    fill_b = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    green  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    teal   = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
    orange = PatternFill(start_color="FBE9E7", end_color="FBE9E7", fill_type="solid")

    for r, rec in enumerate(records, 2):
        base_fill = fill_a if r % 2 == 0 else fill_b
        for c, key in enumerate(all_keys, 1):
            val  = rec.get(key, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            identity_cols = {"display_name","file_name","title","years_exp","education",
                             "edu_degree","edu_field","edu_school","edu_year",
                             "kse_start_year","prev_employers","num_prev_employers","largest_project_value"}
            if val == "Yes":
                if key.startswith("SW_"):   cell.fill = teal
                elif key.startswith("PROJ_"): cell.fill = orange
                else:                         cell.fill = green
            elif val and key not in identity_cols and val not in ("", "No"):
                cell.fill = green
            else:
                cell.fill = base_fill

    # Column widths
    for c, key in enumerate(all_keys, 1):
        if key == "display_name":        ws.column_dimensions[get_column_letter(c)].width = 28
        elif key == "file_name":         ws.column_dimensions[get_column_letter(c)].width = 48
        elif key in ("title","education","edu_school","prev_employers"): ws.column_dimensions[get_column_letter(c)].width = 35
        elif key in ("edu_field","edu_degree","edu_year","kse_start_year","largest_project_value"): ws.column_dimensions[get_column_letter(c)].width = 16
        elif "_num" in key:              ws.column_dimensions[get_column_letter(c)].width = 14
        elif "_exp" in key:              ws.column_dimensions[get_column_letter(c)].width = 10
        else:                            ws.column_dimensions[get_column_letter(c)].width = 9

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_keys))}{len(records)+1}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "KS TalentBridge — Master Index v2 Summary"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Generated:";   ws2["B3"] = NOW
    ws2["A4"] = "Total Staff:"; ws2["B4"] = len(records)
    ws2["A5"] = "Total Columns:"; ws2["B5"] = len(all_keys)
    ws2["A7"] = "Field"; ws2["B7"] = "Staff Count"
    ws2["A7"].font = ws2["B7"].font = Font(bold=True)
    row = 8
    for key in all_keys:
        identity_cols = {"display_name","file_name","title","years_exp","education",
                         "edu_degree","edu_field","edu_school","edu_year",
                         "kse_start_year","prev_employers","num_prev_employers","largest_project_value"}
        if key in identity_cols:
            continue
        count = sum(1 for r in records if r.get(key) and r[key] not in ("", "No"))
        if count > 0:
            ws2.cell(row=row, column=1, value=key.replace("_"," "))
            ws2.cell(row=row, column=2, value=count)
            row += 1
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 15

    wb.save(out_path)
    print(f"  [OK] Excel saved -> {out_path}")

# ── CSV Output ────────────────────────────────────────────────────────────────

def write_csv(records, out_path):
    if not records:
        return
    all_keys = list(dict.fromkeys(k for rec in records for k in rec.keys()))
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=all_keys, extrasaction="ignore")
        writer.writeheader()
        for rec in records:
            row = {k: rec.get(k, "") for k in all_keys}
            writer.writerow(row)
    print(f"  [OK] CSV saved  -> {out_path}")

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    all_files = sorted([
        f for f in os.listdir(RESUME_FOLDER)
        if Path(f).suffix.lower() in RESUME_EXTS
    ])
    total = len(all_files)
    print(f"\nKSE Master Index Builder v2")
    print(f"Scanning {total} resume files...\n")

    raw_records = []
    no_text_files = []
    errors = []

    for i, fname in enumerate(all_files, 1):
        fpath = os.path.join(RESUME_FOLDER, fname)
        print(f"  [{i:>3}/{total}] {fname[:70]}", end="\r")
        try:
            rec, err = build_record(fpath)
            if rec:
                raw_records.append(rec)
            elif err == "no_text":
                no_text_files.append(fname)
        except Exception as e:
            errors.append((fname, str(e)))

    print(" " * 80)
    print(f"\nRaw records extracted : {len(raw_records)}")
    print(f"Files with no text    : {len(no_text_files)}")
    print(f"Errors                : {len(errors)}")

    merged = merge_records(raw_records)
    print(f"Unique staff records  : {len(merged)}\n")

    if errors:
        print("Errors encountered:")
        for f, e in errors[:5]:
            print(f"  - {f}: {e}")

    merged.sort(key=lambda r: r.get("display_name","").lower())

    if merged:
        print(f"Total columns per record: {len(merged[0])}\n")

    print("Writing outputs...")
    xlsx_path = os.path.join(OUT_DIR, "KSE_Master_Index_v2.xlsx")
    csv_path  = os.path.join(OUT_DIR, "KSE_Master_Index_v2.csv")

    write_excel(merged, xlsx_path)
    write_csv(merged, csv_path)

    print(f"\nDone. {len(merged)} staff profiles in {len(merged[0]) if merged else 0} columns.")

if __name__ == "__main__":
    main()
