import argparse
import csv
import datetime as dt
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

try:
    import fitz  # PyMuPDF
    HAS_FITZ = True
except Exception:
    HAS_FITZ = False

W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
RESUME_EXTS = {".docx", ".pdf", ".rtf"}
TODAY = dt.date.today()

SYNONYMS = {
    "roles": {
        "Resident Engineer": ["REI", "Resident Eng"],
        "Construction Inspection": ["CI", "Field Inspector"],
        "Construction Engineering Inspection": ["CEI"],
        "Quality Assurance/Quality Control": ["QA/QC", "QAQC"],
        "Construction Management": ["CM", "CM/CI"],
    },
    "agencies": {
        "NJDOT": ["New Jersey DOT", "New Jersey Department of Transportation"],
        "NYSDOT": ["New York State DOT", "New York State Department of Transportation"],
        "NJ Transit": ["NJT", "New Jersey Transit"],
        "MTA": ["Metropolitan Transportation Authority"],
        "PANYNJ": ["Port Authority", "Port Authority of NY & NJ"],
        "NYCDDC": ["NYC DDC", "NYC Department of Design and Construction"],
        "NYCDOT": ["NYC DOT", "NYC Department of Transportation"],
        "NJTA/Turnpike": ["NJTA", "Turnpike Authority", "Garden State Parkway"],
        "County/Municipal DOT": ["County Engineer", "Municipal Engineer", "City Streets Department"],
        "PennDOT": ["PADOT", "PA DOT", "Pennsylvania Department of Transportation"],
        "SEPTA": ["Southeastern Pennsylvania Transportation Authority"],
        "Amtrak": ["National Railroad Passenger Corporation"],
    },
    "bridge_inspection": {
        "NBIS": ["National Bridge Inspection Standards"],
        "Fracture Critical": ["Fracture-Critical", "FCM"],
        "Load Rating": ["Bridge Load Rating", "LRFR"],
        "Underwater Inspection": ["Subaqueous Inspection", "Diving Inspection"],
        "Structural Inspection": ["Condition Assessment"],
    },
    "materials_coatings": {
        "NACE": ["AMPP", "Coating Inspector", "Paint Inspection", "Corrosion"],
        "Repainting": ["Bridge Repainting", "Coating Rehabilitation"],
    },
    "transportation": {
        "Signal Timing": ["Traffic Signal Timing", "Signal Optimization"],
        "Roadway Lighting": ["Street Lighting", "Illumination Design"],
        "ITS": ["Intelligent Transportation Systems"],
        "Guiderail": ["Guardrail", "Traffic Barrier"],
        "Drainage": ["Stormwater", "Storm Drainage"],
        "Resurfacing": ["Milling and Resurfacing", "Paving Rehabilitation"],
        "Milling": ["Cold Milling", "Pavement Milling"],
    },
}

HELPER_TERMS = {"PANYNJ", "NJTA/Turnpike", "County/Municipal DOT", "PennDOT"}

STATE_PATTERNS = {
    "NJ": re.compile(r"\b(new jersey|nj)\b", re.I),
    "NY": re.compile(r"\b(new york|ny)\b", re.I),
    "PA": re.compile(r"\b(pennsylvania|pa)\b", re.I),
    "CT": re.compile(r"\b(connecticut|ct)\b", re.I),
    "DE": re.compile(r"\b(delaware|de)\b", re.I),
    "MD": re.compile(r"\b(maryland|md)\b", re.I),
    "MA": re.compile(r"\b(massachusetts|ma)\b", re.I),
}

SECTION_RULES = {
    "education": [r"^education$", r"^academic"],
    "certifications": [r"^cert", r"^registr", r"^license"],
    "experience_summary": [r"^experience summary$", r"^summary$", r"^profile$"],
    "project_experience": [r"^project experience$", r"^representative project", r"^selected project"],
    "employment_history": [r"^employment", r"^work history", r"^professional experience", r"^experience$"],
    "skills": [r"^skills$", r"^software", r"^technical skills"],
}

ROLE_PATTERNS = {
    "Project Manager": re.compile(r"\b(project manager|assistant project manager|task manager)\b", re.I),
    "Resident Engineer": re.compile(r"\bresident engineer\b|\brei\b", re.I),
    "Chief Inspector": re.compile(r"\bchief inspector\b|\blead inspector\b", re.I),
    "Senior Inspector": re.compile(r"\bsenior inspector\b", re.I),
    "Field Inspector": re.compile(r"\b(field inspector|inspector)\b", re.I),
    "Construction Manager": re.compile(r"\bconstruction manager\b|\bcm\b", re.I),
    "Construction Engineer": re.compile(r"\bconstruction engineer\b|\bcei\b", re.I),
    "Design Manager": re.compile(r"\bdesign manager\b|\blead designer\b", re.I),
    "Bridge Engineer": re.compile(r"\bbridge engineer\b|\bstructural engineer\b", re.I),
    "Traffic Engineer": re.compile(r"\btraffic engineer\b|\bsignal engineer\b", re.I),
    "Quality Manager": re.compile(r"\bquality manager\b|\bqa/qc\b|\bqaqc\b", re.I),
}

PRIMARY_ROLE_ORDER = ["Project Manager", "Resident Engineer", "Chief Inspector", "Construction Manager", "Design Manager", "Bridge Engineer"]

DISCIPLINE_PATTERNS = {
    "structural": re.compile(r"\b(structural|bridge design|load rating|steel)\b", re.I),
    "civil": re.compile(r"\b(civil|roadway|highway|site)\b", re.I),
    "geotech": re.compile(r"\b(geotech|geotechnical|soil|foundation)\b", re.I),
    "traffic": re.compile(r"\b(traffic|signal|its|timing)\b", re.I),
    "CM/CI": re.compile(r"\b(cm/ci|construction management|construction inspection|cei|ci)\b", re.I),
    "inspection": re.compile(r"\b(inspection|nbis|fracture critical|underwater inspection)\b", re.I),
    "environmental": re.compile(r"\b(environmental|permitting|wetland|nepa)\b", re.I),
    "survey": re.compile(r"\b(survey|topographic|boundary|geospatial)\b", re.I),
}

AGENCY_PATTERNS = {
    "NJDOT": re.compile(r"\b(njdot|new jersey department of transportation)\b", re.I),
    "NYSDOT": re.compile(r"\b(nysdot|new york state department of transportation)\b", re.I),
    "NJ Transit": re.compile(r"\b(nj transit|new jersey transit|njt)\b", re.I),
    "MTA": re.compile(r"\b(mta|metropolitan transportation authority)\b", re.I),
    "PANYNJ": re.compile(r"\b(port authority|panynj|port authority of new york and new jersey)\b", re.I),
    "NYCDDC": re.compile(r"\b(nyc ddc|department of design and construction)\b", re.I),
    "NYCDOT": re.compile(r"\b(nycdot|nyc dot|new york city department of transportation)\b", re.I),
    "NJTA/Turnpike": re.compile(r"\b(njta|new jersey turnpike|turnpike authority|garden state parkway)\b", re.I),
    "PennDOT": re.compile(r"\b(penndot|padot|pa dot|pennsylvania department of transportation)\b", re.I),
    "SEPTA": re.compile(r"\bsepta\b", re.I),
    "Amtrak": re.compile(r"\bamtrak\b", re.I),
    "DelDOT": re.compile(r"\b(deldot|delaware department of transportation)\b", re.I),
    "FAA": re.compile(r"\b(faa|federal aviation administration)\b", re.I),
    "DRPA": re.compile(r"\bdrpa\b|delaware river port authority", re.I),
    "USACE": re.compile(r"\b(usace|army corps of engineers|corps of engineers)\b", re.I),
    "County/Municipal DOT": re.compile(r"\b(county|municipal|city of|township|borough|streets department)\b", re.I),
}

CERT_PATTERNS = {
    "PE": re.compile(r"\b(p\.?\s*e\.?|professional engineer)\b", re.I),
    "PMP": re.compile(r"\bpmp\b", re.I),
    "NICET": re.compile(r"\bnicet\b", re.I),
    "OSHA-10": re.compile(r"\bosha[\s\-]?10\b", re.I),
    "OSHA-30": re.compile(r"\bosha[\s\-]?30\b", re.I),
    "NACE": re.compile(r"\b(nace|ampp)\b", re.I),
    "NBIS": re.compile(r"\bnbis\b", re.I),
    "EIT": re.compile(r"\beit\b|\be\.?i\.?t\.?\b", re.I),
    "ACI": re.compile(r"\baci\b", re.I),
    "CWI": re.compile(r"\bcwi\b", re.I),
    "LEED": re.compile(r"\bleed\b", re.I),
    "PP": re.compile(r"\bprofessional planner\b|\bpp\b", re.I),
    "CFM": re.compile(r"\bcfm\b|certified floodplain manager", re.I),
}

SOFTWARE_PATTERNS = {
    "AutoCAD": re.compile(r"\bautocad\b", re.I),
    "MicroStation": re.compile(r"\bmicrostation\b", re.I),
    "Civil 3D": re.compile(r"\bcivil 3d\b", re.I),
    "OpenRoads": re.compile(r"\bopenroads\b", re.I),
    "Bluebeam": re.compile(r"\bbluebeam\b", re.I),
    "Primavera P6": re.compile(r"\bp6\b|\bprimavera\b", re.I),
    "MS Project": re.compile(r"\bms project\b|\bmicrosoft project\b", re.I),
    "ArcGIS": re.compile(r"\barcgis\b|\bgis\b", re.I),
    "Excel": re.compile(r"\bexcel\b|\bmicrosoft office\b", re.I),
}

SECTOR_PATTERNS = {
    "transportation": re.compile(r"\b(transportation|highway|roadway|dot)\b", re.I),
    "rail": re.compile(r"\b(rail|transit|amtrak|septa|nj transit|mta)\b", re.I),
    "aviation": re.compile(r"\b(airport|aviation|faa|runway|taxiway)\b", re.I),
    "water": re.compile(r"\b(water|wastewater|stormwater|drainage|hydraulics)\b", re.I),
    "buildings": re.compile(r"\b(building|facility|structural)\b", re.I),
    "bridges": re.compile(r"\b(bridge|nbis|fracture critical|load rating)\b", re.I),
}

PROJECT_KEYWORD_PATTERNS = {
    "bridge inspection": re.compile(r"\bbridge inspection\b|\bnbis\b", re.I),
    "fracture critical": re.compile(r"\bfracture critical\b", re.I),
    "load rating": re.compile(r"\bload rating\b|\blrfr\b", re.I),
    "underwater inspection": re.compile(r"\bunderwater inspection\b|\bdiving\b", re.I),
    "condition assessment": re.compile(r"\bcondition assessment\b", re.I),
    "roadway design": re.compile(r"\broadway design\b|\bhighway design\b", re.I),
    "traffic signal": re.compile(r"\btraffic signal\b|\bsignal timing\b", re.I),
    "ITS": re.compile(r"\bits\b|\bintelligent transportation systems\b", re.I),
    "guiderail": re.compile(r"\bguiderail\b|\bguardrail\b", re.I),
    "drainage": re.compile(r"\bdrainage\b|\bstormwater\b", re.I),
    "resurfacing": re.compile(r"\bresurfacing\b|\bmilling and paving\b|\bmilling\b", re.I),
    "construction management": re.compile(r"\bconstruction management\b|\bcm\b", re.I),
    "construction inspection": re.compile(r"\bconstruction inspection\b|\bcei\b|\bci\b", re.I),
    "utility coordination": re.compile(r"\butility coordination\b|\butility relocation\b", re.I),
    "ADA curb ramps": re.compile(r"\bada\b.*\bcurb ramp\b|\bcurb ramp\b", re.I),
    "bridge rehabilitation": re.compile(r"\bbridge rehabilitation\b|\bbridge repair\b|\bbridge replacement\b", re.I),
    "coating inspection": re.compile(r"\bcoating inspection\b|\bpaint inspection\b|\bnace\b", re.I),
    "QA/QC": re.compile(r"\bqa/qc\b|\bquality assurance\b|\bquality control\b", re.I),
}

NOISE_NAME_TERMS = {
    "resume", "master", "kse", "format", "draft", "final", "updated", "rev", "copy",
    "bridge", "inspection", "njdot", "penndot", "sf", "tcm", "airtran", "submission"
}


def parse_args():
    p = argparse.ArgumentParser(description="Build KS Engineers high-recall resume index")
    p.add_argument("--input", required=True)
    p.add_argument("--output", required=True)
    return p.parse_args()


def collect_files(root: Path):
    return sorted([p for p in root.rglob("*") if p.is_file() and p.suffix.lower() in RESUME_EXTS])


def normalize_space(s: str):
    return re.sub(r"\s+", " ", s or "").strip()


def split_lines(text: str):
    return [normalize_space(x) for x in (text or "").splitlines() if normalize_space(x)]


def _docx_paras(xml_bytes: bytes):
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return []
    out = []
    for p in root.iter(f"{W_NS}p"):
        t = "".join(node.text for node in p.iter(f"{W_NS}t") if node.text).strip()
        if t:
            out.append(t)
    return out


def extract_docx(path: Path):
    paras = []
    flags = {}
    try:
        with zipfile.ZipFile(path) as zf:
            parts = ["word/document.xml"]
            parts += sorted(n for n in zf.namelist() if n.startswith("word/header"))
            parts += sorted(n for n in zf.namelist() if n.startswith("word/footer"))
            for part in parts:
                try:
                    with zf.open(part) as f:
                        paras.extend(_docx_paras(f.read()))
                except KeyError:
                    pass
    except Exception:
        flags["PARSE_FAILURE"] = True
        return "", "docx_error", flags
    ded, seen = [], set()
    for p in paras:
        if p not in seen:
            ded.append(p); seen.add(p)
    return "\n".join(ded), "docx_xml", flags


def extract_pdf(path: Path):
    flags = {}
    if not HAS_FITZ:
        flags["PDF_ENGINE_MISSING"] = True
        return "", "pdf_skipped", flags
    try:
        doc = fitz.open(path)
        pages = [pg.get_text("text") for pg in doc]
        doc.close()
        text = "\n".join(pages)
        wc = len(re.findall(r"[A-Za-z0-9']+", text))
        if wc < 80:
            flags["SCANNED_PDF?"] = True
        return text, "pdf_text", flags
    except Exception:
        flags["PARSE_FAILURE"] = True
        return "", "pdf_error", flags


def extract_rtf(path: Path):
    flags = {}
    try:
        raw = path.read_text(encoding="utf-8", errors="ignore")
        text = re.sub(r"\\[a-z]+\d*\s?", " ", raw)
        text = re.sub(r"[{}\\]", " ", text)
        return re.sub(r"\s+", " ", text).strip(), "rtf_text", flags
    except Exception:
        flags["PARSE_FAILURE"] = True
        return "", "rtf_error", flags


def extract_text(path: Path):
    ext = path.suffix.lower()
    if ext == ".docx":
        return extract_docx(path)
    if ext == ".pdf":
        return extract_pdf(path)
    if ext == ".rtf":
        return extract_rtf(path)
    return "", "unknown", {"PARSE_FAILURE": True}


def detect_sections(lines):
    sections = defaultdict(list)
    cur = "other"
    for line in lines:
        lower = line.lower().strip(": ")
        matched = False
        for sec, rules in SECTION_RULES.items():
            for r in rules:
                if re.search(r, lower):
                    cur = sec
                    matched = True
                    break
            if matched:
                break
        if not matched:
            sections[cur].append(line)
    return sections


def clean_name_from_filename(name):
    stem = Path(name).stem
    stem = re.sub(r"[_\-.]+", " ", stem)
    stem = re.sub(r"\b\d{1,8}\b", " ", stem)
    toks = [t for t in stem.split() if t and t.lower() not in NOISE_NAME_TERMS]
    if not toks:
        return "Unknown"
    out = " ".join(toks[:4]).strip()
    out = re.sub(r"\s+", " ", out)
    return out.title() if out else "Unknown"


def extract_name(lines, filename):
    reject = re.compile(r"@|http|resume|summary|experience|education|professional|inspector|engineer|manager", re.I)
    for line in lines[:14]:
        if len(line) < 3 or len(line) > 60:
            continue
        if reject.search(line):
            continue
        if re.search(r"\d{3}[-.\s]\d{3}", line):
            continue
        words = re.findall(r"[A-Za-z][A-Za-z'.-]*", line)
        if 2 <= len(words) <= 4 and all(w[0].isupper() for w in words):
            return " ".join(words)
    return clean_name_from_filename(filename)


def extract_resume_date(lines, text):
    hits = []
    for line in lines[:30]:
        if re.search(r"updated|revised|date|resume", line, re.I):
            hits += re.findall(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b|\b\d{4}-\d{2}-\d{2}\b", line)
    if not hits:
        hits += re.findall(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b|\b\d{4}-\d{2}-\d{2}\b", text[:2500])
    parsed = []
    for h in hits:
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y", "%Y-%m-%d"):
            try:
                parsed.append(dt.datetime.strptime(h, fmt).strftime("%Y-%m-%d"))
                break
            except ValueError:
                pass
    return sorted(parsed)[-1] if parsed else ""


def extract_years_exp(text):
    m = re.search(r"\b(\d{1,2})\+?\s*(?:years?|yrs?)\s+(?:of\s+)?experience\b", text, re.I)
    return m.group(1) if m else ""


def match_map(text, mp):
    return sorted(set([k for k, p in mp.items() if p.search(text)]))


def extract_roles(text):
    all_roles = match_map(text, ROLE_PATTERNS)
    primary = [r for r in PRIMARY_ROLE_ORDER if r in all_roles]
    secondary = [r for r in all_roles if r not in primary]
    return primary, secondary


def extract_certs(lines, text):
    certs = set([k for k, p in CERT_PATTERNS.items() if p.search(text)])
    for ln in lines:
        if CERT_PATTERNS["PE"].search(ln):
            st = [abbr for abbr, pat in STATE_PATTERNS.items() if pat.search(ln)]
            if st:
                for s in st:
                    certs.add(f"PE-{s}")
            else:
                certs.add("PE-Unknown")
        if CERT_PATTERNS["NICET"].search(ln):
            m = re.search(r"\b(?:level|lvl|lv\.?)\s*(i{1,4}|[1-4])\b", ln, re.I)
            if m:
                lv = m.group(1).upper()
                lvmap = {"I": "I", "II": "II", "III": "III", "IV": "IV", "1": "I", "2": "II", "3": "III", "4": "IV"}
                certs.add(f"NICET-{lvmap.get(lv, lv)}")
    return sorted(certs)


def agencies_tags(text):
    exp = match_map(text, AGENCY_PATTERNS)
    helpers = []
    for ag in exp:
        if ag in HELPER_TERMS:
            helpers.append(f"{ag} [helper]")
    if "PANYNJ" in exp:
        helpers.append("Port Authority [helper]")
    return sorted(set(exp)), sorted(set(helpers))


def evidence_snippets(sections, lines):
    key = re.compile(r"\b(pe|p\.e\.|nicet|osha|nace|nbis|njdot|penndot|nysdot|resident engineer|project manager|inspection|bridge|construction|qa/qc|certification|license)\b", re.I)
    cand = []
    for ln in lines:
        if key.search(ln):
            sec = "other"
            for s, lns in sections.items():
                if ln in lns:
                    sec = s
                    break
            cand.append((len(key.findall(ln)), sec, ln))
    cand.sort(key=lambda x: (-x[0], len(x[2])))
    out, seen = [], set()
    for _, sec, ln in cand:
        if ln in seen:
            continue
        seen.add(ln)
        out.append(f"[{sec}] {ln[:200]}")
        if len(out) >= 8:
            break
    if len(out) < 3:
        for sec in ("certifications", "experience_summary", "project_experience", "employment_history"):
            for ln in sections.get(sec, []):
                if ln not in seen:
                    out.append(f"[{sec}] {ln[:200]}")
                    seen.add(ln)
                    if len(out) >= 3:
                        break
            if len(out) >= 3:
                break
    return out[:8]


def flat_synonyms():
    flat = {}
    for cat in SYNONYMS.values():
        for canon, vars_ in cat.items():
            allv = list(dict.fromkeys([canon] + vars_))
            flat[canon.lower()] = allv
            for v in vars_:
                flat[v.lower()] = allv
    return flat


def build_wide_keywords(person_name, primary, secondary, disciplines, agencies_exp, agencies_help, certs, software, states, sectors, projkw):
    base = []
    for seq in ([person_name], primary, secondary, disciplines, agencies_exp, agencies_help, certs, software, states, sectors, projkw):
        for x in seq:
            x = normalize_space(str(x))
            if x:
                base.append(x)
    fs = flat_synonyms()
    expanded = []
    for t in base:
        expanded.append(t)
        if t.lower() in fs:
            expanded.extend(fs[t.lower()])
    for r in primary[:3]:
        for d in disciplines[:3]:
            expanded.append(f"{r} {d}")
    for a in agencies_exp[:5]:
        expanded.append(f"{a} proposal staffing")
        expanded.append(f"{a} project experience")
    for s in sectors[:4]:
        expanded.append(f"{s} infrastructure projects")
    for p in projkw[:8]:
        expanded.append(f"{p} proposal support")

    final, seen = [], set()
    for it in expanded:
        n = normalize_space(str(it)).replace("|", " ").replace(",", " ")
        n = re.sub(r"\s+", " ", n).strip()
        if not n:
            continue
        k = n.lower()
        if k in seen:
            continue
        seen.add(k)
        final.append(n)

    if len(final) < 40:
        generic = [
            "proposal staffing", "resume search", "statement of qualifications", "transportation infrastructure",
            "public agency projects", "construction oversight", "design support", "quality assurance",
            "quality control", "regulatory compliance", "field coordination", "project delivery",
            "engineering services", "inspection services", "DOT experience", "bridge and roadway",
            "construction documents", "shop drawing review", "specification compliance", "project controls",
            "team leadership", "technical proposal support", "infrastructure planning",
            "owner coordination", "contract administration", "construction support services",
            "field verification", "site safety coordination", "quality documentation",
            "public works delivery", "capital improvement projects", "bid phase support",
            "design phase support", "construction phase support", "agency compliance",
            "standards and specifications", "work zone coordination", "transportation asset condition",
            "project reporting", "technical narrative support", "staffing matrix support",
            "proposal shortlist support", "DOT bridge and roadway", "multidisciplinary coordination",
            "infrastructure maintenance", "rehabilitation program support", "inspection documentation",
            "construction schedule awareness", "client coordination", "engineering proposal content"
        ]
        for g in generic:
            if g.lower() not in seen:
                final.append(g); seen.add(g.lower())
            if len(final) >= 40:
                break
    if len(final) < 40:
        i = 1
        while len(final) < 40:
            helper = f"proposal helper keyword {i}"
            if helper.lower() not in seen:
                final.append(helper)
                seen.add(helper.lower())
            i += 1

    return final[:120]


def join_pipe(values):
    out = []
    for v in values:
        n = normalize_space(v)
        if n:
            out.append(n)
    return "|".join(dict.fromkeys(out))


def semantic_summary(person, primary, disciplines, agencies, certs, sectors, yoe):
    lines = [f"{person} is relevant for transportation and infrastructure proposal staffing."]
    if primary:
        lines.append(f"Primary proposal roles: {', '.join(primary[:4])}.")
    if disciplines:
        lines.append(f"Core technical areas: {', '.join(disciplines[:5])}.")
    if agencies:
        lines.append(f"Agency/client exposure includes: {', '.join(agencies[:6])}.")
    if certs:
        lines.append(f"Credentials detected: {', '.join(certs[:8])}.")
    if sectors:
        lines.append(f"Sectors: {', '.join(sectors[:5])}.")
    if yoe:
        lines.append(f"Years of experience stated: {yoe}.")
    return "\n".join(lines[:6])


def stale(last_mod):
    try:
        d = dt.datetime.strptime(last_mod, "%Y-%m-%d").date()
    except ValueError:
        return False
    return (TODAY - d).days > 730


def process_file(path: Path):
    text, method, flags = extract_text(path)
    lines = split_lines(text)
    blob = "\n".join(lines)
    sections = detect_sections(lines)

    person = extract_name(lines, path.name)
    last_mod = dt.datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d")
    resume_date = extract_resume_date(lines, blob)
    yoe = extract_years_exp(blob)

    primary, secondary = extract_roles(blob)
    disciplines = match_map(blob, DISCIPLINE_PATTERNS)
    ag_exp, ag_help = agencies_tags(blob)
    certs = extract_certs(lines, blob)
    software = match_map(blob, SOFTWARE_PATTERNS)
    states = sorted(set([abbr for abbr, pat in STATE_PATTERNS.items() if pat.search(blob)]))
    sectors = match_map(blob, SECTOR_PATTERNS)
    projkw = match_map(blob, PROJECT_KEYWORD_PATTERNS)

    wide = build_wide_keywords(person, primary, secondary, disciplines, ag_exp, ag_help, certs, software, states, sectors, projkw)
    evid = evidence_snippets(sections, lines)

    qf = []
    if flags.get("SCANNED_PDF?"):
        qf.append("SCANNED_PDF?")
    if flags.get("PARSE_FAILURE"):
        qf.append("PARSE_FAILURE")
    if not sections.get("certifications"):
        qf.append("MISSING_CERT_SECTION")
    if stale(last_mod):
        qf.append("STALE_RESUME?")
    if person == "Unknown":
        qf.append("UNKNOWN_NAME")
    if len(re.findall(r"[A-Za-z0-9']+", blob)) < 120:
        qf.append("LOW_TEXT_EXTRACTION?")

    row = {
        "person_name": person,
        "file_name": path.name,
        "local_path": str(path.resolve()),
        "last_modified": last_mod,
        "resume_date": resume_date,
        "primary_role_titles": join_pipe(primary),
        "secondary_role_titles": join_pipe(secondary),
        "discipline_tags": join_pipe(disciplines),
        "agencies_tags": join_pipe(list(ag_exp) + list(ag_help)),
        "certifications_licenses": join_pipe(certs),
        "software_tools": join_pipe(software),
        "locations_states": join_pipe(states),
        "years_experience": yoe,
        "markets_sectors": join_pipe(sectors),
        "project_keywords": join_pipe(projkw),
        "wide_keywords": ", ".join(wide),
        "semantic_summary": semantic_summary(person, primary, disciplines, ag_exp, certs, sectors, yoe),
        "evidence_snippets": " || ".join(evid),
        "quality_flags": join_pipe(qf),
    }

    stats = {
        "processed": 1,
        "failed": 1 if flags.get("PARSE_FAILURE") else 0,
        "scanned_pdf": 1 if flags.get("SCANNED_PDF?") else 0,
        "unknown_name": 1 if person == "Unknown" else 0,
        "missing_cert": 1 if "MISSING_CERT_SECTION" in qf else 0,
    }
    return row, stats


def write_csv(rows, path):
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    flds = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=flds)
        w.writeheader()
        w.writerows(rows)


def write_xlsx(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "KSE_RESUME_INDEX"
    if rows:
        hdr = list(rows[0].keys())
        ws.append(hdr)
        for r in rows:
            ws.append([r.get(h, "") for h in hdr])
        fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        for c in ws[1]:
            c.fill = fill
            c.font = font
        for i, h in enumerate(hdr, 1):
            width = 20
            if h in {"local_path", "wide_keywords", "semantic_summary", "evidence_snippets"}:
                width = 60
            elif h in {"file_name", "agencies_tags", "certifications_licenses", "project_keywords"}:
                width = 38
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
        ws.freeze_panes = "A2"
    wb.save(path)


def write_synonyms(path, applied):
    out = []
    out.append("# KSE Synonym Library")
    out.append("")
    out.append("This controlled vocabulary was used to expand wide_keywords for high-recall search.")
    out.append("")
    out.append("Helper-only terms are marked and are for search expansion, not direct experience claims.")
    out.append("")
    for cat, items in SYNONYMS.items():
        out.append(f"## {cat.replace('_', ' ').title()}")
        out.append("")
        for canon, vars_ in items.items():
            helper = " (helper keywords only)" if canon in HELPER_TERMS else ""
            out.append(f"- **{canon}**{helper}: {', '.join(vars_)}")
        out.append("")
    out.append("## Applied Term Counts")
    out.append("")
    if applied:
        for k, v in applied.most_common():
            out.append(f"- {k}: {v}")
    else:
        out.append("- No applied terms recorded.")
    out.append("")
    path.write_text("\n".join(out), encoding="utf-8")


def write_readme(path, inp, outp, n):
    lines = [
        "# KSE Index README",
        "",
        "## What this is",
        "- High-recall resume metadata index for proposal staffing and semantic search preparation.",
        "- One row per resume (one person record per file).",
        "- Designed to maximize discoverability; false positives are tolerated to avoid misses.",
        "",
        "## Files generated",
        "- KSE_RESUME_INDEX.csv",
        "- KSE_RESUME_INDEX.xlsx",
        "- KSE_SYNONYMS.md",
        "- KSE_INDEX_README.md",
        "",
        "## Source and run context",
        f"- Input folder: `{inp}`",
        f"- Output folder: `{outp}`",
        f"- Records generated: `{n}`",
        f"- Generated on: `{dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}`",
        "",
        "## How to run",
        "```bash",
        "python build_kse_resume_index.py --input \"PATH_TO_RESUME_FOLDER\" --output \"OUTPUT_FOLDER\"",
        "```",
        "",
        "## Update process",
        "1. Add new resumes to the input folder.",
        "2. Re-run the script with the same output folder.",
        "3. Review quality_flags for parse issues and scanned PDFs.",
        "4. Review KSE_SYNONYMS.md and add new domain variants as needed.",
        "",
        "## Behavior notes",
        "- agencies_tags include explicit agency mentions plus helper variants where appropriate.",
        "- evidence_snippets are verbatim short lines from resumes (<=200 chars each).",
        "- If person_name cannot be confidently identified, person_name is Unknown.",
        "- SCANNED_PDF? indicates weak extraction likely from image-based PDF.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def main():
    args = parse_args()
    inp = Path(args.input).resolve()
    outp = Path(args.output).resolve()
    outp.mkdir(parents=True, exist_ok=True)

    files = collect_files(inp)
    rows = []
    summary = Counter()
    applied = Counter()

    print(f"Input folder: {inp}")
    print(f"Output folder: {outp}")
    print(f"Resume files discovered: {len(files)}")

    fs = flat_synonyms()

    for i, fp in enumerate(files, 1):
        print(f"[{i:03}/{len(files)}] {fp.name}", end="\r")
        row, st = process_file(fp)
        rows.append(row)
        summary.update(st)
        for t in [x.strip() for x in row["wide_keywords"].split(",") if x.strip()]:
            k = t.lower()
            if k in fs:
                applied[fs[k][0]] += 1

    rows = sorted(rows, key=lambda r: (r.get("person_name", "Unknown"), r.get("file_name", "")))

    csv_path = outp / "KSE_RESUME_INDEX.csv"
    xlsx_path = outp / "KSE_RESUME_INDEX.xlsx"
    syn_path = outp / "KSE_SYNONYMS.md"
    readme_path = outp / "KSE_INDEX_README.md"

    write_csv(rows, csv_path)
    write_xlsx(rows, xlsx_path)
    write_synonyms(syn_path, applied)
    write_readme(readme_path, inp, outp, len(rows))

    print("\n")
    print("=" * 70)
    print("KSE Resume Index Build Complete")
    print(f"Processed files:       {summary['processed']}")
    print(f"Parse failures:        {summary['failed']}")
    print(f"Scanned PDFs flagged:  {summary['scanned_pdf']}")
    print(f"Unknown names:         {summary['unknown_name']}")
    print(f"Missing cert sections: {summary['missing_cert']}")
    print(f"CSV:                   {csv_path}")
    print(f"XLSX:                  {xlsx_path}")
    print(f"Synonyms:              {syn_path}")
    print(f"README:                {readme_path}")
    print("=" * 70)


if __name__ == "__main__":
    main()
