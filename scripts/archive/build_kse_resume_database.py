"""
build_kse_resume_database.py

Creates a resume "database pack":
1) Raw master table (one row per resume, with full text)
2) Employees table
3) Credentials table
4) Agency experience table
5) Role fit table

Usage:
  python build_kse_resume_database.py --input "PATH_TO_RESUMES" --output "OUTPUT_FOLDER"
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import re
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

try:
    import fitz  # PyMuPDF

    HAS_FITZ = True
except Exception:
    HAS_FITZ = False


W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
RESUME_EXTS = {".docx", ".pdf", ".rtf"}


ROLE_PATTERNS = {
    "Project Manager": re.compile(r"\b(project manager|assistant project manager|task manager)\b", re.I),
    "Resident Engineer": re.compile(r"\bresident engineer\b|\brei\b", re.I),
    "Chief Inspector": re.compile(r"\bchief inspector\b", re.I),
    "Lead Inspector": re.compile(r"\blead inspector\b", re.I),
    "Senior Inspector": re.compile(r"\bsenior inspector\b", re.I),
    "Construction Manager": re.compile(r"\bconstruction manager\b|\bcm\b", re.I),
    "Construction Inspection": re.compile(r"\bconstruction inspection\b|\bci\b|\bcei\b", re.I),
    "Design Manager": re.compile(r"\bdesign manager\b|\blead designer\b", re.I),
    "Bridge Engineer": re.compile(r"\bbridge engineer\b|\bstructural engineer\b", re.I),
    "Traffic Engineer": re.compile(r"\btraffic engineer\b|\bsignal engineer\b", re.I),
    "QA/QC Lead": re.compile(r"\bqa/qc\b|\bquality manager\b|\bquality assurance\b", re.I),
}

DISCIPLINE_PATTERNS = {
    "Bridge Design": re.compile(r"\bbridge design\b|\bstructural design\b", re.I),
    "Bridge Inspection": re.compile(r"\bbridge inspection\b|\bnbis\b|\bfracture critical\b", re.I),
    "Highway Design": re.compile(r"\bhighway design\b|\broadway design\b|\bgeometrics\b", re.I),
    "Traffic Engineering": re.compile(r"\btraffic\b|\bsignal\b|\bits\b", re.I),
    "Hydrology/Hydraulics": re.compile(r"\bhydrology\b|\bhydraulic\b|\bstormwater\b|\bdrainage\b", re.I),
    "CM/CI": re.compile(r"\bconstruction management\b|\bconstruction inspection\b|\bcm/ci\b|\bcei\b", re.I),
    "Geotechnical": re.compile(r"\bgeotechnical\b|\bgeotech\b|\bsoil\b|\bfoundation\b", re.I),
    "Environmental": re.compile(r"\benvironmental\b|\bpermitting\b|\bwetland\b|\bnepa\b", re.I),
    "Steel Rehab": re.compile(r"\bstructural steel repair\b|\bsteel rehabilitation\b|\bsteel repair\b", re.I),
    "Coating QA/QC": re.compile(r"\bcoating inspection\b|\bpaint inspection\b|\bnace\b", re.I),
}

AGENCY_PATTERNS = {
    "NJDOT": re.compile(r"\bnjdot\b|new jersey department of transportation", re.I),
    "PennDOT": re.compile(r"\bpenndot\b|\bpadot\b|pennsylvania department of transportation", re.I),
    "NYSDOT": re.compile(r"\bnysdot\b|new york state department of transportation", re.I),
    "NJ Transit": re.compile(r"\bnj transit\b|\bnjt\b|new jersey transit", re.I),
    "PANYNJ": re.compile(r"\bport authority\b|\bpanynj\b", re.I),
    "SEPTA": re.compile(r"\bsepta\b", re.I),
    "Amtrak": re.compile(r"\bamtrak\b", re.I),
    "DelDOT": re.compile(r"\bdeldot\b|delaware department of transportation", re.I),
    "FAA": re.compile(r"\bfaa\b|federal aviation administration|airport|aviation", re.I),
    "DRPA": re.compile(r"\bdrpa\b|delaware river port authority", re.I),
    "NJDEP": re.compile(r"\bnjdep\b|new jersey department of environmental protection", re.I),
    "USACE": re.compile(r"\busace\b|army corps of engineers|corps of engineers", re.I),
    "Turnpike Authority": re.compile(r"\bnjta\b|new jersey turnpike|turnpike authority|garden state parkway", re.I),
    "City of Newark": re.compile(r"\bcity of newark\b|\bnewark\b", re.I),
    "City of Philadelphia": re.compile(r"\bcity of philadelphia\b|\bphiladelphia\b", re.I),
}

STATE_PATTERNS = {
    "NJ": re.compile(r"\bnew jersey\b|\bnj\b", re.I),
    "PA": re.compile(r"\bpennsylvania\b|\bpa\b", re.I),
    "NY": re.compile(r"\bnew york\b|\bny\b", re.I),
    "CT": re.compile(r"\bconnecticut\b|\bct\b", re.I),
    "DE": re.compile(r"\bdelaware\b|\bde\b", re.I),
    "MD": re.compile(r"\bmaryland\b|\bmd\b", re.I),
}

CREDENTIAL_PATTERNS = {
    "Professional Engineer": re.compile(r"\b(p\.?\s*e\.?|professional engineer)\b", re.I),
    "NICET": re.compile(r"\bnicet\b", re.I),
    "OSHA 10": re.compile(r"\bosha[\s\-]?10\b", re.I),
    "OSHA 30": re.compile(r"\bosha[\s\-]?30\b", re.I),
    "NACE": re.compile(r"\b(nace|ampp)\b", re.I),
    "PMP": re.compile(r"\bpmp\b", re.I),
    "PP": re.compile(r"\bprofessional planner\b|\bpp\b", re.I),
    "CFM": re.compile(r"\bcfm\b|certified floodplain manager", re.I),
    "NBIS": re.compile(r"\bnbis\b", re.I),
    "EIT": re.compile(r"\beit\b|\be\.?i\.?t\.?\b", re.I),
}

SOFTWARE_PATTERNS = {
    "AutoCAD": re.compile(r"\bautocad\b", re.I),
    "MicroStation": re.compile(r"\bmicrostation\b", re.I),
    "Civil 3D": re.compile(r"\bcivil 3d\b", re.I),
    "OpenRoads": re.compile(r"\bopenroads\b", re.I),
    "Bluebeam": re.compile(r"\bbluebeam\b", re.I),
    "Primavera P6": re.compile(r"\bp6\b|\bprimavera\b", re.I),
    "MS Project": re.compile(r"\bms project\b|\bmicrosoft project\b", re.I),
    "ArcGIS": re.compile(r"\barcgis\b|\bgis\b", re.I),
    "Excel": re.compile(r"\bexcel\b|\bmicrosoft office\b", re.I),
}

MARKET_PATTERNS = {
    "Transportation": re.compile(r"\btransportation\b|\bhighway\b|\broadway\b", re.I),
    "Bridges": re.compile(r"\bbridge\b|\bnbis\b", re.I),
    "Rail/Transit": re.compile(r"\brail\b|\btransit\b|\bamtrak\b|\bsepta\b|\bmta\b", re.I),
    "Aviation": re.compile(r"\bairport\b|\baviation\b|\bfaa\b|runway|taxiway", re.I),
    "Water/Utilities": re.compile(r"\bwater\b|\bwastewater\b|\bstormwater\b|\bdrainage\b", re.I),
    "Buildings": re.compile(r"\bbuilding\b|\bfacility\b", re.I),
}

PROJECT_LINE_HINT = re.compile(
    r"\b(project|bridge|roadway|highway|inspection|rehabilitation|replacement|construction|design|improvement|study|corridor|interchange|signal|drainage|streetscape|trail|airport|runway|transit|rail)\b",
    re.I,
)

PROJECT_SECTION_HEADER = re.compile(
    r"^(project experience|relevant project experience|selected projects|representative projects|project assignments|work experience|experience)$",
    re.I,
)

NON_PROJECT_LINE = re.compile(
    r"^(education|certifications|registrations|licenses|professional affiliations|references|skills|software|objective|summary)$",
    re.I,
)


@dataclass
class ExtractResult:
    text: str
    method: str
    scanned_pdf: bool
    parse_failed: bool


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True)
    p.add_argument("--output", required=True)
    return p.parse_args()


def collect_files(root: Path) -> List[Path]:
    return sorted([p for p in root.rglob("*") if p.is_file() and p.suffix.lower() in RESUME_EXTS])


def _docx_paras(xml_bytes: bytes) -> List[str]:
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return []
    out = []
    for p in root.iter(f"{W_NS}p"):
        t = "".join(node.text for node in p.iter(f"{W_NS}t") if node.text).strip()
        if t:
            out.append(t)
    return out


def extract_docx(path: Path) -> ExtractResult:
    paras: List[str] = []
    try:
        with zipfile.ZipFile(path) as zf:
            parts = ["word/document.xml"]
            parts += sorted(n for n in zf.namelist() if n.startswith("word/header"))
            parts += sorted(n for n in zf.namelist() if n.startswith("word/footer"))
            for part in parts:
                try:
                    with zf.open(part) as f:
                        paras.extend(_docx_paras(f.read()))
                except KeyError:
                    pass
        dedup, seen = [], set()
        for p in paras:
            if p not in seen:
                dedup.append(p)
                seen.add(p)
        return ExtractResult("\n".join(dedup), "docx_xml", False, False)
    except Exception:
        return ExtractResult("", "docx_error", False, True)


def extract_pdf(path: Path) -> ExtractResult:
    if not HAS_FITZ:
        return ExtractResult("", "pdf_engine_missing", True, True)
    try:
        doc = fitz.open(path)
        pages = [pg.get_text("text") for pg in doc]
        doc.close()
        text = "\n".join(pages)
        word_count = len(re.findall(r"[A-Za-z0-9']+", text))
        scanned = word_count < 80
        return ExtractResult(text, "pdf_text", scanned, False)
    except Exception:
        return ExtractResult("", "pdf_error", True, True)


def extract_rtf(path: Path) -> ExtractResult:
    try:
        raw = path.read_text(encoding="utf-8", errors="ignore")
        text = re.sub(r"\\[a-z]+\d*\s?", " ", raw)
        text = re.sub(r"[{}\\]", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return ExtractResult(text, "rtf_text", False, False)
    except Exception:
        return ExtractResult("", "rtf_error", False, True)


def extract_text(path: Path) -> ExtractResult:
    ext = path.suffix.lower()
    if ext == ".docx":
        return extract_docx(path)
    if ext == ".pdf":
        return extract_pdf(path)
    if ext == ".rtf":
        return extract_rtf(path)
    return ExtractResult("", "unknown", False, True)


def clean(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip()


def split_lines(text: str) -> List[str]:
    return [clean(x) for x in (text or "").splitlines() if clean(x)]


def extract_name(lines: List[str], file_name: str) -> str:
    reject = re.compile(r"@|http|resume|experience|education|professional|engineer|inspector|manager", re.I)
    for ln in lines[:14]:
        if len(ln) < 3 or len(ln) > 60:
            continue
        if reject.search(ln):
            continue
        words = re.findall(r"[A-Za-z][A-Za-z'.-]*", ln)
        if 2 <= len(words) <= 4 and all(w[0].isupper() for w in words if w):
            return " ".join(words)
    stem = re.sub(r"[_\-.]+", " ", Path(file_name).stem)
    stem = re.sub(r"\b\d{1,8}\b", " ", stem)
    stem = re.sub(r"\b(master|resume|kse|format|draft|final|updated|rev|copy)\b", " ", stem, flags=re.I)
    stem = clean(stem).title()
    return stem if stem else "Unknown"


def find_years_experience(text: str) -> str:
    m = re.search(r"\b(\d{1,2})\+?\s*(?:years?|yrs?)\s+(?:of\s+)?experience\b", text, re.I)
    return m.group(1) if m else ""


def find_resume_date(text: str) -> str:
    hits = re.findall(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b|\b\d{4}-\d{2}-\d{2}\b", text[:4000])
    parsed = []
    for h in hits:
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y", "%Y-%m-%d"):
            try:
                parsed.append(datetime.strptime(h, fmt).strftime("%Y-%m-%d"))
                break
            except ValueError:
                pass
    return sorted(parsed)[-1] if parsed else ""


def match_labels(text: str, patterns: Dict[str, re.Pattern]) -> List[str]:
    return sorted(set([k for k, p in patterns.items() if p.search(text)]))


def evidence_lines(lines: List[str], max_n: int = 6) -> List[str]:
    key = re.compile(r"\b(pe|nicet|osha|nace|nbis|dot|resident engineer|project manager|inspection|bridge|qa/qc)\b", re.I)
    out = []
    seen = set()
    for ln in lines:
        if key.search(ln):
            if ln not in seen:
                out.append(ln[:200])
                seen.add(ln)
            if len(out) >= max_n:
                break
    return out


def _normalize_line_key(line: str) -> str:
    return re.sub(r"\W+", "", (line or "").lower())


def _best_match_from_line(line: str, patterns: Dict[str, re.Pattern]) -> str:
    for label, pattern in patterns.items():
        if pattern.search(line):
            return label
    return ""


def _project_sections(lines: List[str]) -> List[str]:
    selected: List[str] = []
    in_section = False
    for ln in lines:
        l = clean(ln)
        if not l:
            continue
        compact = re.sub(r"[:\- ]+$", "", l.strip())
        if PROJECT_SECTION_HEADER.match(compact):
            in_section = True
            continue
        if in_section and NON_PROJECT_LINE.match(compact):
            in_section = False
        if in_section:
            selected.append(l)
    return selected


def project_lines(lines: List[str]) -> List[str]:
    scoped = _project_sections(lines)
    pool = scoped if scoped else lines
    out: List[str] = []
    seen = set()
    for ln in pool:
        if len(ln) < 20 or len(ln) > 260:
            continue
        if NON_PROJECT_LINE.match(ln):
            continue
        if not PROJECT_LINE_HINT.search(ln) and not _best_match_from_line(ln, AGENCY_PATTERNS):
            continue
        key = _normalize_line_key(ln)
        if key in seen:
            continue
        seen.add(key)
        out.append(ln)
        if len(out) >= 20:
            break
    return out


def project_title_from_line(line: str) -> str:
    chunks = re.split(r"\s[-|:]\s| - | — |\| ", line, maxsplit=1)
    title = clean(chunks[0] if chunks else line)
    if len(title) < 8:
        title = clean(line)
    return title[:120]


def write_csv(path: Path, rows: List[dict]) -> None:
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def write_workbook(path: Path, sheets: Dict[str, List[dict]]) -> None:
    wb = Workbook()
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(name)
        ws.title = name
        first = False
        if not rows:
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        for c in ws[1]:
            c.fill = fill
            c.font = font
    wb.save(path)


def main() -> None:
    args = parse_args()
    input_root = Path(args.input).resolve()
    output_root = Path(args.output).resolve()
    output_root.mkdir(parents=True, exist_ok=True)

    files = collect_files(input_root)
    print(f"Input: {input_root}")
    print(f"Files discovered: {len(files)}")

    raw_rows: List[dict] = []
    employees_by_key: Dict[str, dict] = {}
    creds_rows: List[dict] = []
    agency_rows: List[dict] = []
    role_rows: List[dict] = []
    project_rows: List[dict] = []

    cred_id = 1
    exp_id = 1
    role_id = 1
    project_id = 1
    scanned_pdf_count = 0
    parse_fail_count = 0
    unknown_name_count = 0

    for i, path in enumerate(files, 1):
        print(f"[{i:03}/{len(files)}] {path.name}", end="\r")
        ext = extract_text(path)
        if ext.scanned_pdf:
            scanned_pdf_count += 1
        if ext.parse_failed:
            parse_fail_count += 1

        lines = split_lines(ext.text)
        blob = "\n".join(lines)
        person_name = extract_name(lines, path.name)
        if person_name == "Unknown":
            unknown_name_count += 1

        resume_id = "RES-" + hashlib.md5(str(path).encode("utf-8")).hexdigest()[:10].upper()
        emp_key = re.sub(r"[^a-z0-9]", "", person_name.lower()) or resume_id.lower()
        employee_id = "EMP-" + hashlib.md5(emp_key.encode("utf-8")).hexdigest()[:8].upper()

        last_mod = datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d")
        resume_date = find_resume_date(blob)
        words = len(re.findall(r"[A-Za-z0-9']+", blob))
        roles = match_labels(blob, ROLE_PATTERNS)
        disciplines = match_labels(blob, DISCIPLINE_PATTERNS)
        agencies = match_labels(blob, AGENCY_PATTERNS)
        credentials = match_labels(blob, CREDENTIAL_PATTERNS)
        states = match_labels(blob, STATE_PATTERNS)
        software = match_labels(blob, SOFTWARE_PATTERNS)
        markets = match_labels(blob, MARKET_PATTERNS)
        years_exp = find_years_experience(blob)
        ev = evidence_lines(lines, 6)
        ev_join = " || ".join(ev)

        raw_rows.append(
            {
                "resume_id": resume_id,
                "employee_id": employee_id,
                "person_name": person_name,
                "file_name": path.name,
                "local_path": str(path),
                "file_ext": path.suffix.lower(),
                "last_modified": last_mod,
                "resume_date": resume_date,
                "word_count": words,
                "years_experience": years_exp,
                "roles": "|".join(roles),
                "disciplines": "|".join(disciplines),
                "agencies": "|".join(agencies),
                "credentials": "|".join(credentials),
                "states": "|".join(states),
                "software_tools": "|".join(software),
                "markets_sectors": "|".join(markets),
                "evidence_snippets": ev_join,
                "extraction_method": ext.method,
                "scanned_pdf_flag": "TRUE" if ext.scanned_pdf else "FALSE",
                "parse_failed_flag": "TRUE" if ext.parse_failed else "FALSE",
                "full_text": blob,
            }
        )

        if employee_id not in employees_by_key:
            employees_by_key[employee_id] = {
                "employee_id": employee_id,
                "person_name": person_name,
                "primary_roles": "|".join(roles[:3]),
                "secondary_roles": "|".join(roles[3:]),
                "disciplines": "|".join(disciplines),
                "markets_sectors": "|".join(markets),
                "software_tools": "|".join(software),
                "states": "|".join(states),
                "years_experience": years_exp,
                "resume_count": 0,
                "latest_resume_date": "",
                "latest_last_modified": last_mod,
                "resume_ids": [],
                "resume_paths": [],
            }
        emp = employees_by_key[employee_id]
        emp["resume_count"] += 1
        emp["resume_ids"].append(resume_id)
        emp["resume_paths"].append(str(path))
        if resume_date and (not emp["latest_resume_date"] or resume_date > emp["latest_resume_date"]):
            emp["latest_resume_date"] = resume_date
        if last_mod > emp["latest_last_modified"]:
            emp["latest_last_modified"] = last_mod

        for c in credentials:
            level = ""
            state = ""
            if c == "Professional Engineer":
                for st in states:
                    creds_rows.append(
                        {
                            "credential_id": f"CRD-{cred_id:06d}",
                            "employee_id": employee_id,
                            "credential_type": "Professional Engineer",
                            "credential_name": f"PE-{st}",
                            "level": "",
                            "state": st,
                            "license_number": "",
                            "issue_date": "",
                            "expiration_date": "",
                            "active_status": "Unknown",
                            "evidence_quote": ev[0] if ev else "",
                            "source_resume_id": resume_id,
                            "file_name": path.name,
                            "local_path": str(path),
                        }
                    )
                    cred_id += 1
                if not states:
                    creds_rows.append(
                        {
                            "credential_id": f"CRD-{cred_id:06d}",
                            "employee_id": employee_id,
                            "credential_type": "Professional Engineer",
                            "credential_name": "PE-Unknown",
                            "level": "",
                            "state": "",
                            "license_number": "",
                            "issue_date": "",
                            "expiration_date": "",
                            "active_status": "Unknown",
                            "evidence_quote": ev[0] if ev else "",
                            "source_resume_id": resume_id,
                            "file_name": path.name,
                            "local_path": str(path),
                        }
                    )
                    cred_id += 1
                continue
            if c == "NICET":
                m = re.search(r"\b(?:level|lvl|lv\.?)\s*(I{1,4}|[1-4])\b", blob, re.I)
                level = m.group(1).upper() if m else ""
            creds_rows.append(
                {
                    "credential_id": f"CRD-{cred_id:06d}",
                    "employee_id": employee_id,
                    "credential_type": c,
                    "credential_name": c,
                    "level": level,
                    "state": state,
                    "license_number": "",
                    "issue_date": "",
                    "expiration_date": "",
                    "active_status": "Unknown",
                    "evidence_quote": ev[0] if ev else "",
                    "source_resume_id": resume_id,
                    "file_name": path.name,
                    "local_path": str(path),
                }
            )
            cred_id += 1

        for ag in agencies:
            agency_rows.append(
                {
                    "experience_id": f"EXP-{exp_id:06d}",
                    "employee_id": employee_id,
                    "agency_name": ag,
                    "project_type": disciplines[0] if disciplines else "",
                    "role_context": roles[0] if roles else "",
                    "state_context": states[0] if states else "",
                    "evidence_quote": ev[0] if ev else "",
                    "source_resume_id": resume_id,
                    "file_name": path.name,
                    "local_path": str(path),
                }
            )
            exp_id += 1

        for rl in roles:
            role_rows.append(
                {
                    "role_fit_id": f"RLF-{role_id:06d}",
                    "employee_id": employee_id,
                    "proposal_role": rl,
                    "fit_level": "Medium",
                    "agency_context": agencies[0] if agencies else "",
                    "discipline_context": disciplines[0] if disciplines else "",
                    "evidence_quote": ev[0] if ev else "",
                    "source_resume_id": resume_id,
                    "file_name": path.name,
                    "local_path": str(path),
                }
            )
            role_id += 1

        for ln in project_lines(lines):
            project_rows.append(
                {
                    "project_exp_id": f"PRJ-{project_id:06d}",
                    "employee_id": employee_id,
                    "project_title": project_title_from_line(ln),
                    "agency_client": _best_match_from_line(ln, AGENCY_PATTERNS) or (agencies[0] if agencies else ""),
                    "service_area": _best_match_from_line(ln, DISCIPLINE_PATTERNS) or (disciplines[0] if disciplines else ""),
                    "role_context": _best_match_from_line(ln, ROLE_PATTERNS) or (roles[0] if roles else ""),
                    "market_sector": _best_match_from_line(ln, MARKET_PATTERNS) or (markets[0] if markets else ""),
                    "state_context": _best_match_from_line(ln, STATE_PATTERNS) or (states[0] if states else ""),
                    "evidence_quote": ln[:200],
                    "source_resume_id": resume_id,
                    "file_name": path.name,
                    "local_path": str(path),
                }
            )
            project_id += 1

    employees_rows = []
    for e in employees_by_key.values():
        employees_rows.append(
            {
                "employee_id": e["employee_id"],
                "person_name": e["person_name"],
                "primary_roles": e["primary_roles"],
                "secondary_roles": e["secondary_roles"],
                "disciplines": e["disciplines"],
                "markets_sectors": e["markets_sectors"],
                "software_tools": e["software_tools"],
                "states": e["states"],
                "years_experience": e["years_experience"],
                "resume_count": e["resume_count"],
                "latest_resume_date": e["latest_resume_date"],
                "latest_last_modified": e["latest_last_modified"],
                "resume_ids": "|".join(e["resume_ids"]),
                "resume_paths": "|".join(e["resume_paths"]),
            }
        )

    employees_rows = sorted(employees_rows, key=lambda x: x["person_name"])

    # CSV outputs
    write_csv(output_root / "KSE_Resume_DB_Raw_Master.csv", raw_rows)
    write_csv(output_root / "KSE_Employees.csv", employees_rows)
    write_csv(output_root / "KSE_Credentials.csv", creds_rows)
    write_csv(output_root / "KSE_Agency_Experience.csv", agency_rows)
    write_csv(output_root / "KSE_Role_Fit.csv", role_rows)
    write_csv(output_root / "KSE_Project_Experience.csv", project_rows)

    # Excel workbook with all tables (full_text truncated for Excel cell limits)
    raw_for_excel = []
    for r in raw_rows:
        rx = dict(r)
        full = rx.get("full_text", "")
        rx["full_text"] = full[:32000]
        rx["full_text_excel_truncated"] = "TRUE" if len(full) > 32000 else "FALSE"
        raw_for_excel.append(rx)

    write_workbook(
        output_root / "KSE_Resume_Database.xlsx",
        {
            "Raw_Master": raw_for_excel,
            "Employees": employees_rows,
            "Credentials": creds_rows,
            "Agency_Experience": agency_rows,
            "Role_Fit": role_rows,
            "Project_Experience": project_rows,
        },
    )

    readme = output_root / "KSE_Resume_Database_README.md"
    readme.write_text(
        "\n".join(
            [
                "# KSE Resume Database Pack",
                "",
                "Generated tables:",
                "- KSE_Resume_DB_Raw_Master.csv (one row per resume, full text included)",
                "- KSE_Employees.csv",
                "- KSE_Credentials.csv",
                "- KSE_Agency_Experience.csv",
                "- KSE_Role_Fit.csv",
                "- KSE_Project_Experience.csv",
                "- KSE_Resume_Database.xlsx (all tables in sheets; raw full_text truncated to 32k chars per cell)",
                "",
                "Join key:",
                "- employee_id",
                "",
                "Notes:",
                "- Raw master is the full-recall source table.",
                "- Normalized tables are for deterministic filtering and reporting.",
            ]
        ),
        encoding="utf-8",
    )

    print("\n")
    print("=" * 64)
    print("KSE Resume Database Pack Created")
    print(f"Resumes processed:         {len(raw_rows)}")
    print(f"Unique employees:          {len(employees_rows)}")
    print(f"Credential rows:           {len(creds_rows)}")
    print(f"Agency experience rows:    {len(agency_rows)}")
    print(f"Role fit rows:             {len(role_rows)}")
    print(f"Project experience rows:   {len(project_rows)}")
    print(f"Scanned PDFs flagged:      {scanned_pdf_count}")
    print(f"Parse failures:            {parse_fail_count}")
    print(f"Unknown names:             {unknown_name_count}")
    print(f"Output folder:             {output_root}")
    print("=" * 64)


if __name__ == "__main__":
    main()
